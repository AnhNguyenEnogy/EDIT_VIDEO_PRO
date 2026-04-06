# -*- coding: utf-8 -*-
"""
app_ui.py
=========
Giao diện chính Auto Video Editor — 2 tab: English Vocab + Video Tools.
"""

import sys
import os
import json
import threading
import subprocess as sp
from pathlib import Path
from datetime import datetime

# Bắt exception trong thread — tránh crash app
def _thread_exception_hook(args):
    import traceback
    traceback.print_exception(args.exc_type, args.exc_value, args.exc_traceback)

threading.excepthook = _thread_exception_hook

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QLabel, QPushButton, QLineEdit, QSpinBox, QDoubleSpinBox,
    QComboBox, QGroupBox, QFileDialog, QTextEdit, QCheckBox,
    QSizePolicy, QMessageBox, QScrollArea,
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal
from PyQt6.QtGui import QFont, QFontDatabase, QCursor

PROJECT_ROOT = Path(__file__).parent
FONTS_DIR = PROJECT_ROOT / "fonts"
TOOLS_CONFIG = PROJECT_ROOT / "video_tools_config.json"

# Import PreviewCanvas và TextOverlayEditor từ tab tiếng anh (đã chạy đúng)
from english_vocab.text_overlay_editor import PreviewCanvas, TextOverlayEditor


class FreePreviewCanvas(PreviewCanvas):
    """Mở rộng PreviewCanvas: kéo text tự do cả X và Y."""

    position_xy_changed = pyqtSignal(int, int)  # (new_y_vi, x_offset)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._drag_offset_x = 0

    def _real_to_canvas_x(self, rx):
        r = self._canvas_rect()
        return int(rx / self.canvas_w * r.width()) + r.x()

    def _canvas_to_real_x(self, cx):
        r = self._canvas_rect()
        return int((cx - r.x()) / max(r.width(), 1) * self.canvas_w)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            pos = event.pos()
            if hasattr(self, '_text_rects') and self._text_rects:
                for tr in self._text_rects:
                    if tr and tr.contains(pos):
                        self._dragging = True
                        self._drag_offset_y = pos.y() - tr.center().y()
                        self._drag_offset_x = pos.x() - tr.center().x()
                        self.setCursor(QCursor(Qt.CursorShape.ClosedHandCursor))
                        return

    def mouseMoveEvent(self, event):
        if self._dragging and self.display_data:
            pos = event.pos()
            d = self.display_data
            old_y_vi = d.get("y_vi", 988)
            old_y_en = d.get("y_en", 1058)
            old_y_ipa = d.get("y_ipa", 1130)
            old_x_offset = d.get("x_offset", 0)

            if hasattr(self, '_text_rects') and self._text_rects and self._text_rects[0]:
                # ── Delta Y ──
                old_canvas_y = self._text_rects[0].center().y()
                new_canvas_y = pos.y() - self._drag_offset_y
                delta_canvas_y = new_canvas_y - old_canvas_y
                r = self._canvas_rect()
                delta_real_y = self._canvas_to_real_y(
                    r.y() + delta_canvas_y
                ) - self._canvas_to_real_y(r.y())

                # ── Delta X ──
                old_canvas_x = self._text_rects[0].center().x()
                new_canvas_x = pos.x() - self._drag_offset_x
                delta_canvas_x = new_canvas_x - old_canvas_x
                delta_real_x = self._canvas_to_real_x(
                    r.x() + delta_canvas_x
                ) - self._canvas_to_real_x(r.x())

                new_y_vi = max(0, min(1900, old_y_vi + delta_real_y))
                gap_en = old_y_en - old_y_vi
                gap_ipa = old_y_ipa - old_y_vi
                d["y_vi"] = new_y_vi
                d["y_en"] = new_y_vi + gap_en
                d["y_ipa"] = new_y_vi + gap_ipa

                new_x_offset = old_x_offset + delta_real_x
                new_x_offset = max(-self.canvas_w // 2, min(new_x_offset, self.canvas_w // 2))
                d["x_offset"] = new_x_offset

                self.position_changed.emit(int(new_y_vi))
                self.position_xy_changed.emit(int(new_y_vi), int(new_x_offset))
                self.update()
        else:
            pos = event.pos()
            over = False
            if hasattr(self, '_text_rects'):
                for tr in self._text_rects:
                    if tr and tr.contains(pos):
                        over = True
                        break
            self.setCursor(QCursor(
                Qt.CursorShape.PointingHandCursor if over
                else Qt.CursorShape.OpenHandCursor
            ))

    def _draw_three_lines(self, p, rect, vi, en, ipa, d):
        """Override: thêm x_offset vào vị trí X."""
        from PyQt6.QtGui import QPainterPath, QPen, QBrush, QFontMetrics, QColor
        from PyQt6.QtCore import QRect
        font_name = d.get("font_name", "Arial")
        font_size = d.get("font_size", 62)
        ipa_size = max(font_size - 12, 20)
        bold = d.get("bold", True)
        italic_ipa = True
        text_color = QColor(d.get("text_color", "#FFD700"))
        border_color = QColor(d.get("border_color", "#000000"))
        border_w = d.get("border_width", 4)
        y_vi = d.get("y_vi", 988)
        y_en = d.get("y_en", 1058)
        y_ipa = d.get("y_ipa", 1130)
        x_offset = d.get("x_offset", 0)

        scale = rect.width() / max(self.canvas_w, 1)
        x_offset_canvas = int(x_offset * scale)

        lines_info = [
            (vi,  font_size, bold, False,      y_vi),
            (en,  font_size, bold, False,      y_en),
            (ipa, ipa_size,  False, italic_ipa, y_ipa),
        ]

        self._text_rects = []

        for text, fsize, is_bold, is_italic, real_y in lines_info:
            if not text:
                self._text_rects.append(None)
                continue

            ss = max(1, int(fsize * scale))
            font = QFont(font_name)
            font.setPixelSize(ss)
            font.setBold(is_bold)
            font.setItalic(is_italic)

            fm = QFontMetrics(font)
            tw = fm.horizontalAdvance(text)
            th = fm.height()

            # Canh giữa X + x_offset
            cx = rect.x() + (rect.width() - tw) // 2 + x_offset_canvas
            cy = self._real_to_canvas_y(real_y) + fm.ascent()

            self._text_rects.append(QRect(cx - 4, cy - fm.ascent(), tw + 8, th + 8))

            sbw = max(1, int(border_w * scale))
            path = QPainterPath()
            path.addText(float(cx), float(cy), font, text)

            stroke_pen = QPen(border_color, sbw * 2)
            stroke_pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
            p.setPen(stroke_pen)
            p.setBrush(Qt.BrushStyle.NoBrush)
            p.drawPath(path)

            p.setPen(Qt.PenStyle.NoPen)
            p.setBrush(QBrush(text_color))
            p.drawPath(path)

        if self._dragging:
            for tr in self._text_rects:
                if tr:
                    p.setPen(QPen(QColor(0, 180, 255, 120), 1, Qt.PenStyle.DashLine))
                    p.setBrush(Qt.BrushStyle.NoBrush)
                    p.drawRect(tr)

# ════════════════════════════════════════════════════════════
# CLICKABLE VIDEO WIDGET (click để play/pause)
# ════════════════════════════════════════════════════════════
class ClickableVideoWidget(QWidget):
    """Video widget click để toggle play/pause."""
    clicked = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        from PyQt6.QtMultimediaWidgets import QVideoWidget
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self._video = QVideoWidget()
        self._video.setStyleSheet("background: #080816; border-radius: 8px;")
        layout.addWidget(self._video)
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))

    def video_widget(self):
        return self._video

    def mousePressEvent(self, event):
        self.clicked.emit()
        super().mousePressEvent(event)


# ════════════════════════════════════════════════════════════
# TAB: VIDEO TOOLS (GỘP 3 CHỨC NĂNG)
# ════════════════════════════════════════════════════════════
class VideoToolsTab(QWidget):
    """Tab gộp: Nối video + Phụ đề + Nhạc nền."""

    _sig_done = pyqtSignal(str)  # msg

    def __init__(self, parent=None):
        super().__init__(parent)
        self._sig_done.connect(self._on_done)
        self._running = False
        self._sub_proc = None
        self._loading = False
        self._build_ui()
        self._load_config()

    # ── BUILD UI ──
    def _build_ui(self):
        from PyQt6.QtWidgets import QSplitter, QListWidget
        from PyQt6.QtMultimediaWidgets import QVideoWidget
        from PyQt6.QtMultimedia import QMediaPlayer, QAudioOutput
        from PyQt6.QtCore import QUrl

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        # ════ MAIN SPLITTER: Settings (trái) | Preview (phải) ════
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setObjectName("MainSplitter")

        LBL_W = 90
        SPN_W = 65

        def lbl(text, w=LBL_W):
            l = QLabel(text)
            l.setObjectName("FieldLabel")
            l.setFixedWidth(w)
            return l

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # BÊN TRÁI: SETTINGS (scroll)
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        scroll = QScrollArea()
        scroll.setObjectName("ContentScroll")
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        content = QWidget()
        cl = QVBoxLayout(content)
        cl.setContentsMargins(10, 8, 10, 8)
        cl.setSpacing(6)

        # ── VIDEO NGUỒN + OUTPUT ──
        grp_io = QGroupBox("📂 Video nguồn & Output")
        grp_io.setObjectName("SettingsGroup")
        gio = QVBoxLayout(grp_io)
        gio.setSpacing(4)

        r = QHBoxLayout()
        r.addWidget(lbl("Video:"))
        self.txt_video = QLineEdit()
        self.txt_video.setObjectName("TextInput")
        self.txt_video.setPlaceholderText("Chọn file video hoặc folder...")
        self.txt_video.setReadOnly(True)
        r.addWidget(self.txt_video, 1)
        btn_f = QPushButton("📄 File")
        btn_f.setObjectName("SecondaryButton")
        btn_f.setFixedWidth(65)
        btn_f.setToolTip("Chọn file video")
        btn_f.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_f.clicked.connect(self._browse_video_file)
        r.addWidget(btn_f)
        btn_d = QPushButton("📂 Folder")
        btn_d.setObjectName("SecondaryButton")
        btn_d.setFixedWidth(75)
        btn_d.setToolTip("Chọn folder video")
        btn_d.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_d.clicked.connect(self._browse_video_dir)
        r.addWidget(btn_d)
        gio.addLayout(r)

        r2 = QHBoxLayout()
        r2.addWidget(lbl("Output:"))
        self.txt_output = QLineEdit()
        self.txt_output.setObjectName("TextInput")
        self.txt_output.setPlaceholderText("Folder lưu video output...")
        self.txt_output.setReadOnly(True)
        r2.addWidget(self.txt_output, 1)
        btn_o = QPushButton("📂 Chọn")
        btn_o.setObjectName("SecondaryButton")
        btn_o.setFixedWidth(70)
        btn_o.setToolTip("Chọn folder output")
        btn_o.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_o.clicked.connect(self._browse_output)
        r2.addWidget(btn_o)
        btn_open = QPushButton("📂 Mở")
        btn_open.setObjectName("SecondaryButton")
        btn_open.setFixedWidth(65)
        btn_open.setToolTip("Mở folder output trong Explorer")
        btn_open.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_open.clicked.connect(self._open_output_folder)
        r2.addWidget(btn_open)
        gio.addLayout(r2)

        cl.addWidget(grp_io)

        # ── DỮ LIỆU EXCEL (BATCH) ──
        grp_excel = QGroupBox("📊 Dữ liệu Excel (Tự động hàng loạt)")
        grp_excel.setObjectName("SettingsGroup")
        ge = QVBoxLayout(grp_excel)
        ge.setSpacing(6)
        
        re1 = QHBoxLayout()
        re1.addWidget(QLabel("File Excel:"))
        self.txt_excel = QLineEdit()
        self.txt_excel.setObjectName("FieldInput")
        self.txt_excel.setPlaceholderText("D:/.../tế_công.xlsx")
        re1.addWidget(self.txt_excel)
        btn_sel_excel = QPushButton("📁 Chọn")
        btn_sel_excel.setObjectName("SecondaryButton")
        btn_sel_excel.setFixedWidth(60)
        btn_sel_excel.clicked.connect(self._select_excel)
        re1.addWidget(btn_sel_excel)
        
        btn_check_excel = QPushButton("🔍 Check Data")
        btn_check_excel.setObjectName("PrimaryButton")
        btn_check_excel.clicked.connect(self._check_excel_data)
        re1.addWidget(btn_check_excel)
        ge.addLayout(re1)

        re2 = QHBoxLayout()
        re2.addWidget(QLabel("Cột Prompt:"))
        self.txt_col_prompt = QLineEdit("video prompt")
        self.txt_col_prompt.setFixedWidth(100)
        re2.addWidget(self.txt_col_prompt)
        
        re2.addWidget(QLabel("Cột Upload:"))
        self.txt_col_upload = QLineEdit("video_upload")
        self.txt_col_upload.setFixedWidth(100)
        re2.addWidget(self.txt_col_upload)
        
        re2.addWidget(QLabel("ID Bắt đầu:"))
        self.txt_start_id = QLineEdit("1")
        self.txt_start_id.setFixedWidth(40)
        re2.addWidget(self.txt_start_id)

        re2.addStretch()
        self.chk_excel_enabled = QCheckBox("Bật Batch")
        self.chk_excel_enabled.setObjectName("OptionCheck")
        re2.addWidget(self.chk_excel_enabled)
        ge.addLayout(re2)
        
        cl.addWidget(grp_excel)

        # ── NỐI VIDEO ──
        self.chk_concat = QCheckBox("🔗 Nối Video")
        self.chk_concat.setObjectName("FeatureCheck")
        grp_concat = QGroupBox()
        grp_concat.setObjectName("SettingsGroup")
        gc = QVBoxLayout(grp_concat)
        gc.setSpacing(4)
        gc.addWidget(self.chk_concat)
        rc = QHBoxLayout()
        rc.addWidget(lbl("Video/nhóm:"))
        self.spn_group = QSpinBox()
        self.spn_group.setObjectName("ValueSpin")
        self.spn_group.setRange(1, 50)
        self.spn_group.setValue(4)
        self.spn_group.setFixedWidth(SPN_W)
        rc.addWidget(self.spn_group)
        rc.addSpacing(15)
        self.chk_reencode = QCheckBox("Re-encode")
        self.chk_reencode.setObjectName("OptionCheck")
        self.chk_reencode.setChecked(True)
        self.chk_reencode.setToolTip("Bật nếu video gốc khác codec/resolution.\nTắt = nhanh (chỉ copy stream).")
        rc.addWidget(self.chk_reencode)
        rc.addStretch()
        gc.addLayout(rc)
        cl.addWidget(grp_concat)

        # ── PHỤ ĐỀ TỰ ĐỘNG ──
        self.chk_subtitle = QCheckBox("📝 Phụ đề tự động")
        self.chk_subtitle.setObjectName("FeatureCheck")
        grp_sub = QGroupBox()
        grp_sub.setObjectName("SettingsGroup")
        gs = QVBoxLayout(grp_sub)
        gs.setSpacing(4)
        gs.addWidget(self.chk_subtitle)

        # Dòng 1: Model + Ngôn ngữ
        rs = QHBoxLayout()
        rs.addWidget(lbl("Model:"))
        self.cmb_model = QComboBox()
        self.cmb_model.setObjectName("FontCombo")
        self.cmb_model.addItems(["base", "small", "medium"])
        self.cmb_model.setCurrentText("medium")
        self.cmb_model.setFixedWidth(120)
        self.cmb_model.setToolTip("tiny/base=nhanh kém chính xác\nsmall=vừa\nmedium=tốt nhất cho tiếng Việt\nlarge-v3=chính xác nhất nhưng rất chậm")
        rs.addWidget(self.cmb_model)
        rs.addSpacing(10)
        rs.addWidget(QLabel("🌐"))
        self.cmb_lang = QComboBox()
        self.cmb_lang.setObjectName("FontCombo")
        self.cmb_lang.addItems(["auto", "vi", "en", "ja", "ko", "zh", "fr", "de", "es"])
        self.cmb_lang.setCurrentText("auto")
        self.cmb_lang.setFixedWidth(70)
        rs.addWidget(self.cmb_lang)
        rs.addStretch()
        gs.addLayout(rs)

        
        # Dòng 2: Font + Size + Màu + Nền + Style
        rs2 = QHBoxLayout()
        rs2.addWidget(lbl("Font:"))
        self.cmb_sub_font = QComboBox()
        self.cmb_sub_font.setObjectName("FontCombo")
        self._populate_sub_fonts()
        self.cmb_sub_font.setCurrentText("Arial")
        self.cmb_sub_font.setFixedWidth(180)
        self.cmb_sub_font.setEditable(True)
        rs2.addWidget(self.cmb_sub_font)

        self._hover_font_name = None
        self.cmb_sub_font.highlighted.connect(self._on_font_highlighted)
        self.cmb_sub_font.view().installEventFilter(self)

        rs2.addWidget(QLabel("Size:"))
        self.spn_sub_size = QSpinBox()
        self.spn_sub_size.setObjectName("ValueSpin")
        self.spn_sub_size.setRange(10, 80)
        self.spn_sub_size.setValue(24)
        self.spn_sub_size.setFixedWidth(55)
        rs2.addWidget(self.spn_sub_size)

        # Bold / Italic / Underline toggle buttons
        def _make_style_btn(label, font_weight=None, italic=False, underline=False):
            from PyQt6.QtWidgets import QPushButton
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setFixedSize(26, 26)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            f = QFont("Arial", 9)
            if font_weight: f.setWeight(font_weight)
            if italic: f.setItalic(True)
            if underline: f.setUnderline(True)
            btn.setFont(f)
            btn.setStyleSheet("""
                QPushButton { background: #1e1e3a; color: #ccc; border: 1px solid #444; border-radius: 4px; }
                QPushButton:checked { background: #5a4bd1; color: #fff; border: 1px solid #7c6ff0; }
                QPushButton:hover { background: #2a2a5a; }
            """)
            btn.toggled.connect(lambda: (self._clear_fx_cache() if hasattr(self, '_fx_video_cache') else None, self._update_sub_preview() if hasattr(self, '_fx_video_cache') else None, self._save_config() if hasattr(self, '_fx_video_cache') else None))
            return btn

        self.btn_bold = _make_style_btn("B", font_weight=QFont.Weight.Bold)
        self.btn_bold.setChecked(True)
        rs2.addWidget(self.btn_bold)
        self.btn_italic = _make_style_btn("I", italic=True)
        rs2.addWidget(self.btn_italic)
        self.btn_underline = _make_style_btn("U", underline=True)
        rs2.addWidget(self.btn_underline)

        rs2.addWidget(QLabel("Viền:"))
        self.spn_border_w = QSpinBox()
        self.spn_border_w.setObjectName("ValueSpin")
        self.spn_border_w.setRange(0, 20)
        self.spn_border_w.setValue(3)
        self.spn_border_w.setSuffix("px")
        self.spn_border_w.setFixedWidth(60)
        self.spn_border_w.valueChanged.connect(lambda: (self._clear_fx_cache() if hasattr(self, '_fx_video_cache') else None, self._update_sub_preview() if hasattr(self, '_fx_video_cache') else None, self._save_config() if hasattr(self, '_fx_video_cache') else None))
        rs2.addWidget(self.spn_border_w)

        # Color pickers
        def make_color_btn(color_hex):
            from PyQt6.QtWidgets import QPushButton
            btn = QPushButton()
            btn.setFixedSize(24, 24)
            btn.setStyleSheet(f"background-color: {color_hex}; border: 2px solid #5a4bd1; border-radius: 4px;")
            btn.current_color = color_hex
            return btn

        rs2.addWidget(QLabel("Chữ:"))
        self.btn_text_color = make_color_btn("#FFFFFF")
        self.btn_text_color.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_text_color.clicked.connect(self._pick_text_color)
        rs2.addWidget(self.btn_text_color)

        rs2.addWidget(QLabel("Nền:"))
        self.btn_bg_color = make_color_btn("#000000")
        self.btn_bg_color.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_bg_color.clicked.connect(self._pick_bg_color)
        rs2.addWidget(self.btn_bg_color)

        rs2.addWidget(QLabel("Style:"))
        self.cmb_sub_style = QComboBox()
        self.cmb_sub_style.setObjectName("FontCombo")
        self.cmb_sub_style.addItems([
            "0. Mặc định",
            "1. Classic Yellow",
            "2. White Shadow",
            "3. Fire Red-Orange",
            "4. Cyan Neon Glow",
            "5. Lime Green",
            "6. Pink-Gold",
            "7. Block Background",
            "8. Double Outline",
            "9. Purple Violet",
            "10. Gold Emboss",
            "11. Ice Blue",
            "12. Rainbow",
            "13. Sunset Orange",
            "14. Matrix Green",
            "15. Blood Red",
            "16. Ocean Deep",
            "17. Cotton Candy",
            "18. Chrome Silver",
            "19. Lava Glow",
            "20. Electric Purple",
        ])
        self.cmb_sub_style.setFixedWidth(140)
        self.cmb_sub_style.currentIndexChanged.connect(lambda: (self._clear_fx_cache(), self._update_sub_preview(), self._save_config()))
        rs2.addWidget(self.cmb_sub_style)

        rs2.addStretch()
        gs.addLayout(rs2)

        # Dòng 2.5: Karaoke Colors
        rs_kara = QHBoxLayout()
        rs_kara.addWidget(lbl("Kara Chạy:"))
        self.btn_kara_run_color = make_color_btn("#FFFF00") # Mặc định vàng
        self.btn_kara_run_color.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_kara_run_color.clicked.connect(self._pick_kara_run_color)
        rs_kara.addWidget(self.btn_kara_run_color)

        rs_kara.addWidget(lbl("Kara Nền:", 70))
        self.btn_kara_bg_color = make_color_btn("#FFFFFF") # Mặc định trắng
        self.btn_kara_bg_color.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_kara_bg_color.clicked.connect(self._pick_kara_bg_color)
        rs_kara.addWidget(self.btn_kara_bg_color)
        
        rs_kara.addStretch()
        gs.addLayout(rs_kara)


        # Dòng 3: Vị trí + Căn chỉnh + Hiệu ứng
        rs3 = QHBoxLayout()
        rs3.addWidget(lbl("Vị trí:"))
        self.cmb_sub_pos = QComboBox()
        self.cmb_sub_pos.setObjectName("FontCombo")
        self.cmb_sub_pos.addItems(["⬇ Dưới", "⬆ Trên", "↔ Giữa"])
        self.cmb_sub_pos.setCurrentIndex(0)
        self.cmb_sub_pos.setFixedWidth(80)
        rs3.addWidget(self.cmb_sub_pos)
        rs3.addWidget(QLabel("Căn:"))
        self.cmb_sub_align = QComboBox()
        self.cmb_sub_align.setObjectName("FontCombo")
        self.cmb_sub_align.addItems(["⇔ Giữa", "⇐ Trái", "⇒ Phải"])
        self.cmb_sub_align.setCurrentIndex(0)
        self.cmb_sub_align.setFixedWidth(80)
        rs3.addWidget(self.cmb_sub_align)
        rs3.addWidget(QLabel("Khung:"))
        self.cmb_aspect = QComboBox()
        self.cmb_aspect.setObjectName("FontCombo")
        self.cmb_aspect.addItems(["📱 9:16", "🖥 16:9"])
        self.cmb_aspect.setCurrentIndex(0)
        self.cmb_aspect.setFixedWidth(80)
        self.cmb_aspect.setToolTip("Video dọc (9:16) hoặc ngang (16:9)")
        rs3.addWidget(self.cmb_aspect)
        rs3.addStretch()
        gs.addLayout(rs3)

        # Dòng 4: FX + MarginV
        rs4 = QHBoxLayout()
        rs4.addWidget(lbl("FX:"))
        self.cmb_sub_fx = QComboBox()
        self.cmb_sub_fx.setObjectName("FontCombo")
        self.cmb_sub_fx.addItems([
            "❌ Không",
            "🎬 Fade in/out",
            "💡 Highlight từng từ",
            "🔥 Word Pop (CapCut)",
            "🎆 Pháo hoa biến mất",
            "💬 Pop-up",
            "⌨️ Typewriter",
            "🏀 Bounce",
            "⬅️ Slide trái",
            "➡️ Slide phải",
            "✨ Glow neon",
        ])
        self.cmb_sub_fx.setCurrentIndex(0)
        self.cmb_sub_fx.setFixedWidth(200)
        self.cmb_sub_fx.setToolTip(
            "Không: phụ đề bình thường\n"
            "Fade in/out: xuất hiện/biến mất mờ dần\n"
            "Highlight từng từ: đọc đến đâu sáng đến đó (karaoke)\n"
            "Word Pop (CapCut): từng từ hiện lên với POP effect, từ đang đọc nổi bật\n"
            "Pháo hoa biến mất: chữ đọc xong bung nổ phóng to + bay mất\n"
            "Pop-up: text nhỏ → to\n"
            "Typewriter: hiện từng từ một\n"
            "Bounce: chữ nảy lên\n"
            "Slide trái/phải: text trượt vào từ mép\n"
            "Glow neon: text phát sáng"
        )
        rs4.addWidget(self.cmb_sub_fx)

        # ── Hover preview FX: cache + debounce timer ──
        self._fx_video_cache = {}     # {fx_name: Path}
        self._fx_cache_dir = None     # thư mục tạm chứa video cache
        self._fx_pending_idx = -1
        self._fx_hover_timer = QTimer()
        self._fx_hover_timer.setSingleShot(True)
        self._fx_hover_timer.setInterval(350)
        self._fx_hover_timer.timeout.connect(self._on_fx_hover_timeout)
        self.cmb_sub_fx.highlighted.connect(self._on_fx_highlighted)
        rs4.addSpacing(10)
        rs4.addWidget(QLabel("Margin:"))
        self.spn_margin_v = QSpinBox()
        self.spn_margin_v.setObjectName("ValueSpin")
        self.spn_margin_v.setRange(0, 1200)
        self.spn_margin_v.setValue(30)
        self.spn_margin_v.setSuffix("px")
        self.spn_margin_v.setFixedWidth(70)
        self.spn_margin_v.setToolTip("Khoảng cách text từ mép (px).\nKéo text trên preview để chỉnh.")
        rs4.addWidget(self.spn_margin_v)
        rs4.addStretch()
        gs.addLayout(rs4)

        cl.addWidget(grp_sub)

        # ── NHẠC NỀN (gọn 2 dòng) ──
        self.chk_music = QCheckBox("🎵 Nhạc nền")
        self.chk_music.setObjectName("FeatureCheck")
        grp_music = QGroupBox()
        grp_music.setObjectName("SettingsGroup")
        gm = QVBoxLayout(grp_music)
        gm.setSpacing(4)
        gm.addWidget(self.chk_music)

        rm1 = QHBoxLayout()
        rm1.addWidget(lbl("Nhạc:"))
        self.txt_music = QLineEdit()
        self.txt_music.setObjectName("TextInput")
        self.txt_music.setPlaceholderText("File nhạc...")
        self.txt_music.setReadOnly(True)
        rm1.addWidget(self.txt_music, 1)
        btn_m = QPushButton("🎵 Chọn")
        btn_m.setObjectName("SecondaryButton")
        btn_m.setFixedWidth(70)
        btn_m.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_m.clicked.connect(self._browse_music)
        rm1.addWidget(btn_m)
        gm.addLayout(rm1)

        # Gộp tất cả settings nhạc trên 1 dòng
        rm2 = QHBoxLayout()
        rm2.addWidget(lbl("Vol:", 30))
        self.spn_volume = QDoubleSpinBox()
        self.spn_volume.setObjectName("ValueSpin")
        self.spn_volume.setRange(0.0, 1.0)
        self.spn_volume.setValue(0.3)
        self.spn_volume.setSingleStep(0.05)
        self.spn_volume.setDecimals(2)
        self.spn_volume.setFixedWidth(55)
        rm2.addWidget(self.spn_volume)
        rm2.addWidget(QLabel("Từ:"))
        self.spn_m_start = QDoubleSpinBox()
        self.spn_m_start.setObjectName("ValueSpin")
        self.spn_m_start.setRange(0, 9999)
        self.spn_m_start.setValue(0)
        self.spn_m_start.setSuffix("s")
        self.spn_m_start.setFixedWidth(60)
        rm2.addWidget(self.spn_m_start)
        rm2.addWidget(QLabel("→"))
        self.spn_m_end = QDoubleSpinBox()
        self.spn_m_end.setObjectName("ValueSpin")
        self.spn_m_end.setRange(0, 9999)
        self.spn_m_end.setValue(0)
        self.spn_m_end.setSuffix("s")
        self.spn_m_end.setToolTip("0 = tự cắt theo video")
        self.spn_m_end.setFixedWidth(60)
        rm2.addWidget(self.spn_m_end)
        rm2.addWidget(QLabel("FI:"))
        self.spn_fade_in = QDoubleSpinBox()
        self.spn_fade_in.setObjectName("ValueSpin")
        self.spn_fade_in.setRange(0, 30)
        self.spn_fade_in.setValue(1.0)
        self.spn_fade_in.setSuffix("s")
        self.spn_fade_in.setFixedWidth(50)
        rm2.addWidget(self.spn_fade_in)
        rm2.addWidget(QLabel("FO:"))
        self.spn_fade_out = QDoubleSpinBox()
        self.spn_fade_out.setObjectName("ValueSpin")
        self.spn_fade_out.setRange(0, 30)
        self.spn_fade_out.setValue(2.0)
        self.spn_fade_out.setSuffix("s")
        self.spn_fade_out.setFixedWidth(50)
        rm2.addWidget(self.spn_fade_out)
        self.btn_preview = QPushButton("🔊 Nghe thử")
        self.btn_preview.setObjectName("SecondaryButton")
        self.btn_preview.setFixedWidth(85)
        self.btn_preview.setToolTip("Nghe thử 10 giây")
        self.btn_preview.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._preview_playing = False
        self.btn_preview.clicked.connect(self._toggle_preview)
        rm2.addWidget(self.btn_preview)
        rm2.addStretch()
        gm.addLayout(rm2)

        cl.addWidget(grp_music)

        # ── NÚT CHẠY (toggle) ──
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.btn_run = QPushButton("🚀 Chạy")
        self.btn_run.setObjectName("PrimaryButton")
        self.btn_run.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_run.clicked.connect(self._toggle_run)
        self.btn_run.setFixedWidth(140)
        btn_row.addWidget(self.btn_run)
        btn_row.addStretch()
        cl.addLayout(btn_row)
        # ── LOG ──
        self.txt_log = QTextEdit()
        self.txt_log.setObjectName("LogPanel")
        self.txt_log.setReadOnly(True)
        self.txt_log.setFont(QFont("Consolas", 10))
        self.txt_log.setPlaceholderText("Thông tin log sẽ hiển thị ở đây...")

        # Splitter dọc bên trái: Settings (trên) | Log (dưới)
        left_splitter = QSplitter(Qt.Orientation.Vertical)
        left_splitter.setObjectName("LeftSplitter")
        
        scroll.setWidget(content)
        left_splitter.addWidget(scroll)
        left_splitter.addWidget(self.txt_log)
        
        left_splitter.setStretchFactor(0, 4)
        left_splitter.setStretchFactor(1, 1)

        splitter.addWidget(left_splitter)


        
        

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # BÊN PHẢI: VIDEO PREVIEW
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        preview_widget = QWidget()
        preview_widget.setObjectName("PreviewPanel")
        pv_layout = QVBoxLayout(preview_widget)
        pv_layout.setContentsMargins(6, 8, 10, 8)
        pv_layout.setSpacing(6)

        pv_header = QHBoxLayout()
        pv_title = QLabel("🎬 Kết quả")
        pv_title.setObjectName("FieldLabel")
        pv_title.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
        pv_header.addWidget(pv_title)
        pv_header.addStretch()
        btn_refresh = QPushButton("🔄 Làm mới")
        btn_refresh.setObjectName("SecondaryButton")
        btn_refresh.setFixedWidth(85)
        btn_refresh.setToolTip("Làm mới danh sách file output")
        btn_refresh.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_refresh.clicked.connect(self._refresh_output_list)
        pv_header.addWidget(btn_refresh)
        pv_layout.addLayout(pv_header)

        # Preview subtitle — FreePreviewCanvas: kéo X + Y tự do
        self._sub_preview = FreePreviewCanvas()
        self._sub_preview.position_xy_changed.connect(self._on_preview_drag_xy)
        self._current_x_offset = 0
        self._y_vi_from_drag = None
        pv_layout.addWidget(self._sub_preview, 1)

        # Video player — tạo lazy, chỉ khi phát video
        self._clickable_video = None
        self._pv_layout = pv_layout  # lưu ref để addWidget sau
        
        # UI cho phần Excel (Batch) sẽ thêm sau


        self._player = QMediaPlayer()
        self._audio_output = QAudioOutput()
        self._player.setAudioOutput(self._audio_output)

        # Chuyển page khi phát/dừng video
        from PyQt6.QtMultimedia import QMediaPlayer as _QMP
        self._player.playbackStateChanged.connect(self._on_player_state)

        # Thanh tua + thời gian
        from PyQt6.QtWidgets import QSlider
        seek_row = QHBoxLayout()
        self.lbl_time_current = QLabel("00:00")
        self.lbl_time_current.setObjectName("FieldLabel")
        self.lbl_time_current.setFixedWidth(45)
        seek_row.addWidget(self.lbl_time_current)

        self.slider_seek = QSlider(Qt.Orientation.Horizontal)
        self.slider_seek.setObjectName("SeekSlider")
        self.slider_seek.setRange(0, 0)
        self.slider_seek.sliderMoved.connect(self._seek_video)
        self._slider_pressed = False
        self.slider_seek.sliderPressed.connect(lambda: setattr(self, '_slider_pressed', True))
        self.slider_seek.sliderReleased.connect(self._on_slider_released)
        seek_row.addWidget(self.slider_seek, 1)

        self.lbl_time_total = QLabel("00:00")
        self.lbl_time_total.setObjectName("FieldLabel")
        self.lbl_time_total.setFixedWidth(45)
        seek_row.addWidget(self.lbl_time_total)

        btn_open_sys = QPushButton("🖥️ Mở player")
        btn_open_sys.setObjectName("SecondaryButton")
        btn_open_sys.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_open_sys.clicked.connect(self._open_with_system)
        seek_row.addWidget(btn_open_sys)
        pv_layout.addLayout(seek_row)

        # Danh sách file output
        self.lst_files = QListWidget()
        self.lst_files.setObjectName("FileList")
        self.lst_files.setMaximumHeight(150)
        self.lst_files.itemClicked.connect(self._play_selected)
        pv_layout.addWidget(self.lst_files)

        splitter.addWidget(preview_widget)

        # Tỉ lệ splitter: settings 55% | preview 45%
        splitter.setSizes([550, 450])
        splitter.setStretchFactor(0, 5)
        splitter.setStretchFactor(1, 4)

        layout.addWidget(splitter)

        # Timer cập nhật thanh tua + thời gian
        self._timer = QTimer()
        self._timer.setInterval(250)
        self._timer.timeout.connect(self._update_player_time)
        self._timer.start()

        # Kết nối signals cập nhật preview khi đổi settings
        self.cmb_sub_font.currentTextChanged.connect(lambda: (
            setattr(self, '_hover_font_name', None),
            self._clear_fx_cache(), self._update_sub_preview(), self._save_config()
        ))
        self.spn_sub_size.valueChanged.connect(lambda: (self._clear_fx_cache(), self._update_sub_preview(), self._save_config()))
        self.cmb_sub_style.currentIndexChanged.connect(lambda: (self._clear_fx_cache(), self._update_sub_preview(), self._save_config()))
        self.cmb_sub_pos.currentIndexChanged.connect(self._on_preset_changed)
        self.cmb_sub_align.currentIndexChanged.connect(self._on_preset_changed)
        self.cmb_sub_fx.currentIndexChanged.connect(lambda: (self._update_sub_preview(), self._save_config()))
        self.cmb_aspect.currentIndexChanged.connect(self._on_preset_changed)
        self.spn_margin_v.valueChanged.connect(self._on_margin_changed)

        # Auto-save khi thay đổi bất kỳ setting nào
        self.chk_concat.toggled.connect(lambda: self._save_config())
        self.spn_group.valueChanged.connect(lambda: self._save_config())
        self.chk_reencode.toggled.connect(lambda: self._save_config())
        self.chk_subtitle.toggled.connect(lambda: self._save_config())
        self.cmb_model.currentIndexChanged.connect(lambda: self._save_config())
        self.cmb_lang.currentIndexChanged.connect(lambda: self._save_config())
        self.chk_music.toggled.connect(lambda: self._save_config())
        self.spn_volume.valueChanged.connect(lambda: self._save_config())
        self.spn_m_start.valueChanged.connect(lambda: self._save_config())
        self.spn_m_end.valueChanged.connect(lambda: self._save_config())
        self.spn_fade_in.valueChanged.connect(lambda: self._save_config())
        self.spn_fade_out.valueChanged.connect(lambda: self._save_config())
        self.chk_excel_enabled.toggled.connect(lambda: self._save_config())
        self.txt_col_prompt.textChanged.connect(lambda: self._save_config())
        self.txt_col_upload.textChanged.connect(lambda: self._save_config())
        self.txt_start_id.textChanged.connect(lambda: self._save_config())

        # Vẽ preview lần đầu (delay nhỏ để widget có kích thước)
        QTimer.singleShot(200, self._update_sub_preview)

    def _populate_sub_fonts(self):
        """Quét font hệ thống + thư mục fonts/ (giống English Vocab)."""
        from PyQt6.QtGui import QFontDatabase, QFont, QFontMetrics
        fonts_set = set()
        for family in QFontDatabase.families():
            fonts_set.add(family)
        if FONTS_DIR.exists():
            for f in FONTS_DIR.glob("*.ttf"):
                fid = QFontDatabase.addApplicationFont(str(f))
                if fid >= 0:
                    for fam in QFontDatabase.applicationFontFamilies(fid):
                        fonts_set.add(fam)
            for f in FONTS_DIR.glob("*.otf"):
                fid = QFontDatabase.addApplicationFont(str(f))
                if fid >= 0:
                    for fam in QFontDatabase.applicationFontFamilies(fid):
                        fonts_set.add(fam)

        # Lọc font hỗ trợ tiếng Việt: kiểm tra ký tự đặc trưng
        viet_test = "ăâêôơưđ"
        vi_fonts = set()
        for fam in fonts_set:
            try:
                fm = QFontMetrics(QFont(fam, 12))
                if all(fm.horizontalAdvance(ch) > 0 for ch in viet_test):
                    vi_fonts.add(fam)
            except Exception:
                pass

        priority = [
            "Arial", "Be Vietnam Pro", "Roboto", "Montserrat", "Poppins",
            "Nunito", "Open Sans", "Lato", "Quicksand",
            "Segoe UI", "Tahoma", "Times New Roman", "Verdana", "Calibri",
        ]
        result = []
        for p in priority:
            if p in vi_fonts:
                result.append(p)
                vi_fonts.discard(p)
        result.extend(sorted(vi_fonts))
        self.cmb_sub_font.addItems(result)

    def _pick_text_color(self):
        from PyQt6.QtWidgets import QColorDialog
        from PyQt6.QtGui import QColor
        init_col = QColor(self.btn_text_color.current_color)
        col = QColorDialog.getColor(init_col, self, "Chọn màu chữ")
        if col.isValid():
            hex_c = col.name().upper()
            self.btn_text_color.current_color = hex_c
            self.btn_text_color.setStyleSheet(f"background-color: {hex_c}; border: 2px solid #5a4bd1; border-radius: 4px;")
            self._clear_fx_cache()
            self._update_sub_preview()
            self._save_config()

    def _pick_bg_color(self):
        from PyQt6.QtWidgets import QColorDialog
        from PyQt6.QtGui import QColor
        init_col = QColor(self.btn_bg_color.current_color)
        col = QColorDialog.getColor(init_col, self, "Chọn màu nền/viền")
        if col.isValid():
            hex_c = col.name().upper()
            self.btn_bg_color.current_color = hex_c
            self.btn_bg_color.setStyleSheet(f"background-color: {hex_c}; border: 2px solid #5a4bd1; border-radius: 4px;")
            self._clear_fx_cache()
            self._update_sub_preview()
            self._save_config()

    def _pick_kara_run_color(self):
        from PyQt6.QtWidgets import QColorDialog
        from PyQt6.QtGui import QColor
        init_col = QColor(self.btn_kara_run_color.current_color)
        col = QColorDialog.getColor(init_col, self, "Chọn màu chữ karaoke chạy")
        if col.isValid():
            hex_c = col.name().upper()
            self.btn_kara_run_color.current_color = hex_c
            self.btn_kara_run_color.setStyleSheet(f"background-color: {hex_c}; border: 2px solid #5a4bd1; border-radius: 4px;")
            self._clear_fx_cache()
            self._update_sub_preview()
            self._save_config()

    def _pick_kara_bg_color(self):
        from PyQt6.QtWidgets import QColorDialog
        from PyQt6.QtGui import QColor
        init_col = QColor(self.btn_kara_bg_color.current_color)
        col = QColorDialog.getColor(init_col, self, "Chọn màu chữ karaoke nền")
        if col.isValid():
            hex_c = col.name().upper()
            self.btn_kara_bg_color.current_color = hex_c
            self.btn_kara_bg_color.setStyleSheet(f"background-color: {hex_c}; border: 2px solid #5a4bd1; border-radius: 4px;")
            self._clear_fx_cache()
            self._update_sub_preview()
            self._save_config()

    def _on_margin_changed(self):
        """Khi user chỉnh margin tay → reset y_vi_from_drag."""
        self._y_vi_from_drag = None
        self._update_sub_preview()
        self._save_config()

    def _on_font_highlighted(self, index):
        """Khi hover qua font trong dropdown → preview font đó ngay."""
        font_name = self.cmb_sub_font.itemText(index)
        if font_name:
            self._hover_font_name = font_name
            self._update_sub_preview()

    def eventFilter(self, obj, event):
        """Detect popup font đóng → revert preview về font đã chọn."""
        if obj == self.cmb_sub_font.view() and event.type() == event.Type.Hide:
            if self._hover_font_name:
                self._hover_font_name = None
                self._update_sub_preview()
        return super().eventFilter(obj, event)

    def _on_preset_changed(self):
        """Khi user đổi preset vị trí → reset y_vi_from_drag và x_offset."""
        self._y_vi_from_drag = None
        self._current_x_offset = 0
        self._update_sub_preview()
        self._save_config()

    # ── PREVIEW HIỆU ỨNG FX (HOVER TỰ ĐỘNG) ──
    _FX_MAP = {0: "none", 1: "fade", 2: "karaoke", 3: "word_pop",
               4: "firework", 5: "popup", 6: "typewriter", 7: "bounce",
               8: "slide_left", 9: "slide_right", 10: "glow"}

    def _on_fx_highlighted(self, index):
        """Khi hover qua item FX trong dropdown → debounce và play preview."""
        self._fx_pending_idx = index
        self._fx_hover_timer.start()  # Reset timer mỗi lần hover sang item mới

    def _on_fx_hover_timeout(self):
        """Debounce xong → tạo hoặc phát video preview FX."""
        idx = self._fx_pending_idx
        fx = self._FX_MAP.get(idx, "none")

        if fx == "none":
            # Dừng video nếu chọn "Không"
            if self._player.playbackState() != self._player.PlaybackState.StoppedState:
                self._player.stop()
            return

        # Kiểm tra cache
        if fx in self._fx_video_cache:
            cached = self._fx_video_cache[fx]
            if cached.exists():
                self._play_fx_cached(cached)
                return

        # Chưa có cache → tạo video
        video_path = self._build_fx_video(fx)
        if video_path and video_path.exists():
            self._fx_video_cache[fx] = video_path
            self._play_fx_cached(video_path)

    def _play_fx_cached(self, video_path):
        """Phát video FX từ cache."""
        self._ensure_video_widget()
        from PyQt6.QtCore import QUrl
        self._player.setSource(QUrl.fromLocalFile(str(video_path)))
        self._player.play()

    def _clear_fx_cache(self):
        """Xóa cache khi settings thay đổi (font/size/color...)."""
        self._fx_video_cache.clear()
        # Cleanup thư mục tạm cũ
        if self._fx_cache_dir and self._fx_cache_dir.exists():
            import shutil
            try:
                shutil.rmtree(self._fx_cache_dir, ignore_errors=True)
            except Exception:
                pass
        self._fx_cache_dir = None

    def _get_fx_cache_dir(self):
        """Lấy (hoặc tạo) thư mục cache."""
        if self._fx_cache_dir is None or not self._fx_cache_dir.exists():
            import tempfile
            self._fx_cache_dir = Path(tempfile.mkdtemp(prefix="fx_cache_"))
        return self._fx_cache_dir

    def _build_fx_video(self, fx):
        """Tạo video preview cho 1 hiệu ứng FX. Trả về Path video."""
        try:
            from ffmpeg_utils import get_ffmpeg
            from subtitle_generator import SubtitleGenerator

            aspect_idx = self.cmb_aspect.currentIndex()
            if aspect_idx == 0:
                vw, vh = 720, 1280
            else:
                vw, vh = 1280, 720

            font_name = self.cmb_sub_font.currentText()
            font_size = int(self.spn_sub_size.value() * 1.33)
            font_color = getattr(self, "btn_text_color", type("o", (), {"current_color": "#FFFFFF"})).current_color
            bg_color = getattr(self, "btn_bg_color", type("o", (), {"current_color": "#000000"})).current_color
            kara_run_color = getattr(self, "btn_kara_run_color", type("o", (), {"current_color": "#FFFF00"})).current_color
            kara_bg_color = getattr(self, "btn_kara_bg_color", type("o", (), {"current_color": "#FFFFFF"})).current_color
            style_idx = getattr(self, "cmb_sub_style", type("o", (), {"currentIndex": lambda: 0})).currentIndex()

            margin_v = self.spn_margin_v.value()
            _pos_map = {0: "bottom", 1: "top", 2: "center"}
            _align_map_ass = {0: 2, 1: 1, 2: 3}
            position = _pos_map.get(self.cmb_sub_pos.currentIndex(), "bottom")
            alignment = _align_map_ass.get(self.cmb_sub_align.currentIndex(), 2)

            cache_dir = self._get_fx_cache_dir()
            ffmpeg = get_ffmpeg()

            # 1. Tạo video nền (dùng chung cho tất cả fx)
            blank_video = cache_dir / "blank.mp4"
            if not blank_video.exists():
                cmd_blank = [
                    ffmpeg, "-y",
                    "-f", "lavfi",
                    "-i", f"color=c=0x0A0A1A:s={vw}x{vh}:d=4:r=30",
                    "-c:v", "libx264", "-preset", "ultrafast",
                    "-pix_fmt", "yuv420p",
                    str(blank_video),
                ]
                sp.run(cmd_blank, capture_output=True, creationflags=0x08000000)

            # 2. Segments mẫu với word timestamps
            segments = [
                {"start": 0.0, "end": 1.8, "text": "Xin chào thế giới", "words": [
                    {"word": "Xin", "start": 0.0, "end": 0.3},
                    {"word": "chào", "start": 0.3, "end": 0.6},
                    {"word": "thế", "start": 0.6, "end": 0.9},
                    {"word": "giới", "start": 0.9, "end": 1.5},
                ]},
                {"start": 2.0, "end": 3.8, "text": "Hiệu ứng phụ đề", "words": [
                    {"word": "Hiệu", "start": 2.0, "end": 2.3},
                    {"word": "ứng", "start": 2.3, "end": 2.6},
                    {"word": "phụ", "start": 2.6, "end": 2.9},
                    {"word": "đề", "start": 2.9, "end": 3.5},
                ]},
            ]

            # 3. Tạo ASS
            ass_path = cache_dir / f"{fx}.ass"
            sg = SubtitleGenerator(log_callback=lambda m: None)
            sg._write_ass(
                ass_path, segments,
                font_name=font_name, font_size=font_size,
                font_color=font_color, bg_color=bg_color, style_idx=style_idx, alignment=alignment,
                position=position, fx=fx,
                margin_v=margin_v, aspect_idx=aspect_idx,
                kara_run_color=kara_run_color, kara_bg_color=kara_bg_color,
            )

            # 4. Burn
            output = cache_dir / f"{fx}.mp4"
            ass_str = str(ass_path).replace("\\", "/").replace(":", "\\:")
            fonts_dir = str(FONTS_DIR).replace("\\", "/").replace(":", "\\:")
            if FONTS_DIR.exists():
                vf = f"ass='{ass_str}':fontsdir='{fonts_dir}'"
            else:
                vf = f"ass='{ass_str}'"

            cmd_burn = [
                ffmpeg, "-y",
                "-i", str(blank_video),
                "-vf", vf,
                "-c:v", "libx264", "-preset", "ultrafast",
                "-pix_fmt", "yuv420p",
                str(output),
            ]
            result = sp.run(cmd_burn, capture_output=True, text=True,
                           encoding="utf-8", errors="replace", creationflags=0x08000000)

            if result.returncode != 0:
                return None

            return output if output.exists() else None

        except Exception as e:
            import traceback
            traceback.print_exc()
            return None

    def _calc_y_vi(self):
        """Tính y_vi giống hệt preview — dùng cho burn."""
        if getattr(self, '_y_vi_from_drag', None) is not None:
            return self._y_vi_from_drag

        aspect_idx = self.cmb_aspect.currentIndex()
        vh = 1280 if aspect_idx == 0 else 720
        font_size = self.spn_sub_size.value()
        margin_v = self.spn_margin_v.value()
        pos_idx = self.cmb_sub_pos.currentIndex()

        if pos_idx == 0:      # bottom
            return vh - margin_v - font_size * 3
        elif pos_idx == 1:    # top
            return margin_v
        else:                 # center
            return (vh - font_size * 3) // 2

    # ── PREVIEW PHỤ ĐỀ ──
    def _update_sub_preview(self):
        """Cập nhật display_data cho PreviewCanvas."""
        aspect_idx = self.cmb_aspect.currentIndex()
        if aspect_idx == 0:
            vw, vh = 720, 1280
        else:
            vw, vh = 1280, 720

        # Không ghi đè nếu đang drag
        if self._sub_preview._dragging:
            return

        self._sub_preview.canvas_w = vw
        self._sub_preview.canvas_h = vh

        margin_v = self.spn_margin_v.value()
        font_size = self.spn_sub_size.value()
        pos_idx = self.cmb_sub_pos.currentIndex()

        # Nếu đã drag → dùng y_vi đã lưu, chỉ tính lại khi chưa drag hoặc margin thay đổi
        if not hasattr(self, '_y_vi_from_drag') or self._y_vi_from_drag is None:
            if pos_idx == 0:
                y_vi = vh - margin_v - font_size * 3
            elif pos_idx == 1:
                y_vi = margin_v
            else:
                y_vi = (vh - font_size * 3) // 2
        else:
            y_vi = self._y_vi_from_drag

        y_en = y_vi + int(font_size * 1.4)

        tc = getattr(self, "btn_text_color", type("o", (), {"current_color": "#FFFFFF"})).current_color
        bc = getattr(self, "btn_bg_color", type("o", (), {"current_color": "#000000"})).current_color
        style_idx = getattr(self, "cmb_sub_style", type("o", (), {"currentIndex": lambda: 0})).currentIndex()

        # Nếu là Karaoke, preview cái màu chữ chạy
        is_kara = self.cmb_sub_fx.currentIndex() in (2, 3) # karaoke hoặc word_pop
        if is_kara:
            tc = getattr(self, "btn_kara_run_color", type("o", (), {"current_color": "#FFFF00"})).current_color
            bc = getattr(self, "btn_bg_color", type("o", (), {"current_color": "#000000"})).current_color # Mặc định viền đen cho preview

        self._sub_preview.set_display_data({
            "word1": {
                "vi": "Người ta sống vì hạnh phúc",
                "en": "và những khoảnh khắc đẹp",
                "ipa": "",
            },
            "font_name": self._hover_font_name or self.cmb_sub_font.currentText(),
            "font_size": font_size,
            "bold": self.btn_bold.isChecked(),
            "italic": self.btn_italic.isChecked(),
            "underline": self.btn_underline.isChecked(),
            "text_color": tc,
            "border_color": bc,
            "border_width": self.spn_border_w.value(),
            "style_idx": style_idx,
            "y_vi": y_vi,
            "y_en": y_en,
            "y_ipa": y_en + int(font_size * 1.2),
            "x_offset": self._current_x_offset,
        })

    def _on_preview_drag_xy(self, new_y_vi, x_offset):
        """Khi kéo text trên canvas → lưu y_vi, x_offset, cập nhật margin spinbox."""
        self._current_x_offset = x_offset
        self._y_vi_from_drag = new_y_vi

        font_size = self.spn_sub_size.value()
        pos_idx = self.cmb_sub_pos.currentIndex()
        aspect_idx = self.cmb_aspect.currentIndex()
        vh = 1280 if aspect_idx == 0 else 720

        if pos_idx == 0:
            margin = vh - new_y_vi - font_size * 3
        elif pos_idx == 1:
            margin = new_y_vi
        else:
            margin = vh - new_y_vi - font_size * 3

        margin = max(0, min(int(margin), 1200))
        self.spn_margin_v.blockSignals(True)
        self.spn_margin_v.setValue(margin)
        self.spn_margin_v.blockSignals(False)
        self._save_config()  # TỰ ĐỘNG LƯU PHÁT MỘT

    def _ensure_video_widget(self):
        """Tạo ClickableVideoWidget lần đầu khi cần phát video."""
        if self._clickable_video is None:
            self._clickable_video = ClickableVideoWidget()
            self._clickable_video.clicked.connect(self._toggle_play)
            # Insert ngay sau header, trước canvas
            self._pv_layout.insertWidget(1, self._clickable_video)
            self._player.setVideoOutput(self._clickable_video.video_widget())

    def _on_player_state(self, state):
        """Show/hide video player khi phát/dừng."""
        from PyQt6.QtMultimedia import QMediaPlayer as _QMP
        if self._clickable_video is None:
            return
        if state == _QMP.PlaybackState.StoppedState:
            self._clickable_video.hide()
            self._sub_preview.show()
        else:
            self._sub_preview.hide()
            self._clickable_video.show()

    # ── HELPERS ──
    @staticmethod
    def _close_excel_if_open(filepath):
        """Đóng file Excel trong ứng dụng Excel (không lưu) nếu đang mở."""
        try:
            import win32com.client
            xl = win32com.client.GetObject(Class="Excel.Application")
        except Exception:
            return False  # Excel không chạy
        try:
            target = str(Path(filepath).resolve()).lower()
            for wb in xl.Workbooks:
                if str(Path(wb.FullName).resolve()).lower() == target:
                    wb.Close(SaveChanges=False)
                    return True
        except Exception:
            pass
        return False

    def _log(self, msg):
        from PyQt6.QtCore import QMetaObject, Qt as QtNS, Q_ARG
        ts = datetime.now().strftime("%H:%M:%S")
        QMetaObject.invokeMethod(
            self.txt_log, "append",
            QtNS.ConnectionType.QueuedConnection,
            Q_ARG(str, f"[{ts}] {msg}"))

    def _log_raw(self, msg):
        from PyQt6.QtCore import QMetaObject, Qt as QtNS, Q_ARG
        QMetaObject.invokeMethod(
            self.txt_log, "append",
            QtNS.ConnectionType.QueuedConnection,
            Q_ARG(str, msg))

    # ── BROWSE ──
    def _browse_video_file(self):
        p, _ = QFileDialog.getOpenFileName(
            self, "Chọn video", "", "Video (*.mp4 *.mkv *.avi *.mov);;All (*)")
        if p:
            self.txt_video.setText(p)
            self._save_config()

    def _browse_video_dir(self):
        p = QFileDialog.getExistingDirectory(self, "Chọn folder video")
        if p:
            self.txt_video.setText(p)
            self._save_config()

    def _browse_output(self):
        p = QFileDialog.getExistingDirectory(self, "Chọn folder output")
        if p:
            self.txt_output.setText(p)
            self._save_config()

    def _browse_music(self):
        p, _ = QFileDialog.getOpenFileName(
            self, "Chọn nhạc", "", "Audio (*.mp3 *.wav *.m4a *.aac *.ogg);;All (*)")
        if p:
            self.txt_music.setText(p)
            self._save_config()

    # ── NHẠC PREVIEW ──
    def _preview_music(self):
        music = self.txt_music.text().strip()
        if not music or not Path(music).exists():
            QMessageBox.warning(self, "Thiếu", "Chọn file nhạc trước!")
            return
        from background_music import BackgroundMusic
        if not hasattr(self, '_bgm'):
            self._bgm = BackgroundMusic(log_callback=self._log)
        self._bgm.preview_music(music, start_time=self.spn_m_start.value(), duration=10,
                                volume=self.spn_volume.value())

    def _stop_preview(self):
        if hasattr(self, '_bgm'):
            self._bgm.stop_preview()

    def _toggle_preview(self):
        """Toggle nghe thử / dừng nghe."""
        if self._preview_playing:
            self._stop_preview()
            self._preview_playing = False
            self.btn_preview.setText("🔊 Nghe thử")
            self.btn_preview.setObjectName("SecondaryButton")
        else:
            self._preview_music()
            self._preview_playing = True
            self.btn_preview.setText("⏹ Dừng nghe")
            self.btn_preview.setObjectName("DangerButton")
        self.btn_preview.style().unpolish(self.btn_preview)
        self.btn_preview.style().polish(self.btn_preview)

    def _check_excel_data(self):
        """Kiểm tra file Excel (ID tịnh tiến cho mỗi cảnh của hàng trống)."""
        self._save_config()
        path_excel = self.txt_excel.text()
        video_dir = Path(self.txt_video.text())
        start_id_val = int(self.txt_start_id.text() if self.txt_start_id.text().isdigit() else 1)
        
        if not path_excel or not Path(path_excel).exists():
            self._log("❌ Chưa chọn file Excel.")
            return
        if not video_dir.exists():
            self._log("❌ Thư mục video gốc không tồn tại.")
            return

        self._log(f"🔍 ĐANG CHECK DATA (Bắt đầu ID: {start_id_val})...")
        try:
            import openpyxl, re
            self._close_excel_if_open(path_excel)
            wb = openpyxl.load_workbook(path_excel, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() for c in next(ws.rows)]
            
            cpr = headers.index(self.txt_col_prompt.text().lower()) if self.txt_col_prompt.text().lower() in headers else -1
            cup = headers.index(self.txt_col_upload.text().lower()) if self.txt_col_upload.text().lower() in headers else -1

            if cpr == -1 or cup == -1:
                self._log(f"❌ Lỗi: Cột '{self.txt_col_prompt.text()}' hoặc '{self.txt_col_upload.text()}' không có.")
                return

            valid_count = 0
            current_id = start_id_val 

            for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                prompt = row[cpr].value
                v_upload = row[cup].value
                if not prompt: continue
                
                is_empty = not v_upload or not str(v_upload).strip()
                if not is_empty:
                    # Bỏ qua hàng đã xử lý
                    continue

                nsc = len([line for line in str(prompt).splitlines() if line.strip()])
                # Các ID cho hàng này: current_id -> current_id + nsc - 1
                row_ids = list(range(current_id, current_id + nsc))
                self._log(f"📋 Hàng {ridx} | Cảnh: {nsc} | Cần Video ID: {row_ids}")
                
                ms = []
                for vid in row_ids:
                    # Tìm theo pattern: id_1_*.mp4
                    pat = f"{vid}_1_*.mp4"
                    fnd = list(video_dir.glob(pat))
                    if not fnd:
                        # Thử fnd vid_1.mp4 
                        fnd = list(video_dir.glob(f"{vid}_1.mp4"))
                    
                    if not fnd: ms.append(vid)
                
                if ms: 
                    self._log(f"   ⚠️ Thiếu Video ID: {ms}")
                else: 
                    valid_count += 1
                    self._log(f"   ✅ Đủ (OK)")
                
                # Quan trọng: Tăng ID cho hàng tiếp theo (Kể cả hàng hiện tại thiếu file cũng tăng ID để hàng sau ko bị đè)
                current_id += nsc

            self._log(f"📊 Sẵn sàng: {valid_count} hàng.")
            wb.close()
        except Exception as e:
            self._log(f"❌ Lỗi khi check: {e}")

    def _toggle_run(self):
        """Toggle chạy / dừng."""
        if self._running:
            self._stop_all()
        else:
            self._run_all()

    # ── VIDEO PLAYER + FOLDER ──
    def _open_output_folder(self):
        out = self.txt_output.text().strip()
        if out and Path(out).exists():
            os.startfile(out)
        else:
            QMessageBox.information(self, "Thông báo", "Folder output chưa có hoặc chưa chọn.")

    def _refresh_output_list(self):
        self.lst_files.clear()
        out = self.txt_output.text().strip()
        if not out or not Path(out).exists():
            return
        mp4s = sorted(Path(out).glob("*.mp4"))
        for f in mp4s:
            self.lst_files.addItem(f.name)

    def _select_excel(self):
        f, _ = QFileDialog.getOpenFileName(self, "Chọn file Excel", "", "Excel Files (*.xlsx *.xls)")
        if f:
            self.txt_excel.setText(f)
            self._save_config()

    def _play_selected(self, item):
        from PyQt6.QtCore import QUrl
        out = self.txt_output.text().strip()
        if not out:
            return
        fpath = Path(out) / item.text()
        if fpath.exists():
            self._ensure_video_widget()
            self._player.setSource(QUrl.fromLocalFile(str(fpath)))
            self._player.play()

    def _toggle_play(self):
        from PyQt6.QtMultimedia import QMediaPlayer
        state = self._player.playbackState()
        if state == QMediaPlayer.PlaybackState.PlayingState:
            self._player.pause()
        elif state == QMediaPlayer.PlaybackState.PausedState:
            self._player.play()
        # Nếu stopped và có source → play lại
        elif self._player.source().isValid():
            self._player.play()

    def _seek_video(self, position):
        """Khi kéo thanh tua."""
        self._player.setPosition(position)
        # Cập nhật label thời gian ngay
        pos_s = position // 1000
        self.lbl_time_current.setText(f"{pos_s//60:02d}:{pos_s%60:02d}")

    def _on_slider_released(self):
        self._slider_pressed = False
        self._player.setPosition(self.slider_seek.value())

    def _open_with_system(self):
        """Mở video đang chọn bằng player mặc định hệ thống."""
        item = self.lst_files.currentItem()
        if not item:
            QMessageBox.information(self, "Thông báo", "Chọn video trong danh sách trước.")
            return
        out = self.txt_output.text().strip()
        if out:
            fpath = Path(out) / item.text()
            if fpath.exists():
                os.startfile(str(fpath))

    def _update_player_time(self):
        from PyQt6.QtMultimedia import QMediaPlayer
        if self._player.playbackState() != QMediaPlayer.PlaybackState.StoppedState:
            pos = self._player.position()
            dur = self._player.duration()
            pos_s = pos // 1000
            dur_s = dur // 1000

            self.lbl_time_current.setText(f"{pos_s//60:02d}:{pos_s%60:02d}")
            self.lbl_time_total.setText(f"{dur_s//60:02d}:{dur_s%60:02d}")

            # Cập nhật slider (nếu user không đang kéo)
            if not self._slider_pressed:
                if self.slider_seek.maximum() != dur:
                    self.slider_seek.setRange(0, dur)
                self.slider_seek.setValue(pos)

    # ════════════════════════════════════════════════════════
    # SAVE / LOAD CONFIG
    # ════════════════════════════════════════════════════════
    def _save_config(self):
        if self._loading:
            return
        try:
            cfg = {
                "video": self.txt_video.text(),
                "output": self.txt_output.text(),
                "concat_enabled": self.chk_concat.isChecked(),
                "group_size": self.spn_group.value(),
                "reencode": self.chk_reencode.isChecked(),
                "subtitle_enabled": self.chk_subtitle.isChecked(),
                "model": self.cmb_model.currentText(),
                "language": self.cmb_lang.currentText(),
                "sub_font": self.cmb_sub_font.currentText(),
                "sub_size": self.spn_sub_size.value(),
                "sub_color_hex": getattr(self, 'btn_text_color', type('o', (), {'current_color': '#FFFFFF'})).current_color,
                "sub_bg_color_hex": getattr(self, 'btn_bg_color', type('o', (), {'current_color': '#000000'})).current_color,
                "kara_run_color_hex": getattr(self, 'btn_kara_run_color', type('o', (), {'current_color': '#FFFF00'})).current_color,
                "kara_bg_color_hex": getattr(self, 'btn_kara_bg_color', type('o', (), {'current_color': '#FFFFFF'})).current_color,
                "sub_style_idx": getattr(self.cmb_sub_style, 'currentIndex', lambda: 0)(),
                "sub_bold": getattr(self, 'btn_bold', type('o', (), {'isChecked': lambda: True})).isChecked(),
                "sub_italic": getattr(self, 'btn_italic', type('o', (), {'isChecked': lambda: False})).isChecked(),
                "sub_underline": getattr(self, 'btn_underline', type('o', (), {'isChecked': lambda: False})).isChecked(),
                "sub_border_w": getattr(self, 'spn_border_w', type('o', (), {'value': lambda: 3})).value(),
                "sub_pos_idx": self.cmb_sub_pos.currentIndex(),
                "sub_align_idx": self.cmb_sub_align.currentIndex(),
                "sub_fx_idx": self.cmb_sub_fx.currentIndex(),
                "aspect_idx": self.cmb_aspect.currentIndex(),
                "margin_v": self.spn_margin_v.value(),
                "sub_x_offset": getattr(self, '_current_x_offset', 0),
                "sub_y_vi": getattr(self, '_y_vi_from_drag', None),
                "music_enabled": self.chk_music.isChecked(),
                "music_path": self.txt_music.text(),
                "music_volume": self.spn_volume.value(),
                "music_start": self.spn_m_start.value(),
                "music_end": self.spn_m_end.value(),
                "fade_in": self.spn_fade_in.value(),
                "fade_out": self.spn_fade_out.value(),
                "excel_path": self.txt_excel.text(),
                "excel_enabled": self.chk_excel_enabled.isChecked(),
                "col_prompt": self.txt_col_prompt.text(),
                "col_upload": self.txt_col_upload.text(),
                "start_id": str(self.txt_start_id.text() if self.txt_start_id.text().isdigit() else '1'),
            }
            TOOLS_CONFIG.write_text(
                json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass

    def _load_config(self):
        if not TOOLS_CONFIG.exists():
            return
        self._loading = True
        try:
            cfg = json.loads(TOOLS_CONFIG.read_text(encoding="utf-8"))
            if cfg.get("video"):
                self.txt_video.setText(cfg["video"])
            if cfg.get("output"):
                self.txt_output.setText(cfg["output"])
            self.chk_concat.setChecked(cfg.get("concat_enabled", False))
            self.spn_group.setValue(cfg.get("group_size", 4))
            self.chk_reencode.setChecked(cfg.get("reencode", True))
            self.chk_subtitle.setChecked(cfg.get("subtitle_enabled", False))
            idx = self.cmb_model.findText(cfg.get("model", "medium"))
            if idx >= 0:
                self.cmb_model.setCurrentIndex(idx)
            idx = self.cmb_lang.findText(cfg.get("language", "auto"))
            if idx >= 0:
                self.cmb_lang.setCurrentIndex(idx)
            # Subtitle style
            if cfg.get("sub_font"):
                self.cmb_sub_font.setCurrentText(cfg["sub_font"])
            self.spn_sub_size.setValue(cfg.get("sub_size", 24))
            if hasattr(self, 'btn_text_color'):
                c = cfg.get("sub_color_hex", "#FFFFFF")
                self.btn_text_color.current_color = c
                self.btn_text_color.setStyleSheet(f"background-color: {c}; border: 2px solid #5a4bd1; border-radius: 4px;")
            if hasattr(self, 'btn_bg_color'):
                c = cfg.get("sub_bg_color_hex", "#000000")
                self.btn_bg_color.current_color = c
                self.btn_bg_color.setStyleSheet(f"background-color: {c}; border: 2px solid #5a4bd1; border-radius: 4px;")
            if hasattr(self, 'btn_kara_run_color'):
                c = cfg.get("kara_run_color_hex", "#FFFF00")
                self.btn_kara_run_color.current_color = c
                self.btn_kara_run_color.setStyleSheet(f"background-color: {c}; border: 2px solid #5a4bd1; border-radius: 4px;")
            if hasattr(self, 'btn_kara_bg_color'):
                c = cfg.get("kara_bg_color_hex", "#FFFFFF")
                self.btn_kara_bg_color.current_color = c
                self.btn_kara_bg_color.setStyleSheet(f"background-color: {c}; border: 2px solid #5a4bd1; border-radius: 4px;")
            if hasattr(self, 'cmb_sub_style'):
                self.cmb_sub_style.setCurrentIndex(cfg.get("sub_style_idx", 0))
            if hasattr(self, 'btn_bold'):
                self.btn_bold.setChecked(cfg.get("sub_bold", True))
            if hasattr(self, 'btn_italic'):
                self.btn_italic.setChecked(cfg.get("sub_italic", False))
            if hasattr(self, 'btn_underline'):
                self.btn_underline.setChecked(cfg.get("sub_underline", False))
            if hasattr(self, 'spn_border_w'):
                self.spn_border_w.setValue(cfg.get("sub_border_w", 3))
            self.cmb_sub_pos.setCurrentIndex(cfg.get("sub_pos_idx", 0))
            self.cmb_sub_align.setCurrentIndex(cfg.get("sub_align_idx", 0))
            self.cmb_sub_fx.setCurrentIndex(cfg.get("sub_fx_idx", 0))
            self.cmb_aspect.setCurrentIndex(cfg.get("aspect_idx", 0))
            self.spn_margin_v.setValue(cfg.get("margin_v", 30))
            self._current_x_offset = cfg.get("sub_x_offset", 0)
            self._y_vi_from_drag = cfg.get("sub_y_vi", None)
            
            self.txt_excel.setText(cfg.get("excel_path", "") or cfg.get("batch_excel", ""))
            
            self.txt_col_prompt.setText(cfg.get("col_prompt", "video prompt"))
            
            self.txt_col_upload.setText(cfg.get("col_upload", "video_upload"))
            self.txt_start_id.setText(str(cfg.get("start_id", "1")))
            self.chk_excel_enabled.setChecked(cfg.get("excel_enabled", False))

            self.chk_music.setChecked(cfg.get("music_enabled", False))
            if cfg.get("music_path"):
                self.txt_music.setText(cfg["music_path"])
            self.spn_volume.setValue(cfg.get("music_volume", 0.3))
            self.spn_m_start.setValue(cfg.get("music_start", 0))
            self.spn_m_end.setValue(cfg.get("music_end", 0))
            self.spn_fade_in.setValue(cfg.get("fade_in", 1.0))
            self.spn_fade_out.setValue(cfg.get("fade_out", 2.0))
        except Exception:
            pass
        finally:
            self._loading = False

    # ════════════════════════════════════════════════════════
    # RUN ALL (theo checkbox)
    # ════════════════════════════════════════════════════════
    def _run_all(self):
        video = self.txt_video.text().strip()
        if not video or not Path(video).exists():
            QMessageBox.warning(self, "Thiếu", "Chọn video nguồn trước!")
            return

        do_concat = self.chk_concat.isChecked()
        do_sub = self.chk_subtitle.isChecked()
        do_music = self.chk_music.isChecked()

        if not do_concat and not do_sub and not do_music:
            QMessageBox.warning(self, "Thiếu", "Chọn ít nhất 1 chức năng để chạy!")
            return

        output_dir = self.txt_output.text().strip()
        if not output_dir:
            if Path(video).is_file():
                output_dir = str(Path(video).parent / "output")
            else:
                output_dir = str(Path(video) / "output")
            self.txt_output.setText(output_dir)

        self._save_config()
        self._running = True
        self.btn_run.setText("⏹ Dừng xử lý")
        self.btn_run.setObjectName("DangerButton")
        self.btn_run.style().unpolish(self.btn_run)
        self.btn_run.style().polish(self.btn_run)

        # Mapping màu → ASS color format (&HBBGGRR)
        _color_map = {
            0: "&HFFFFFF",   # Trắng
            1: "&H00FFFF",   # Vàng
            2: "&HFF6400",   # Xanh dương
            3: "&H00FF00",   # Xanh lá
            4: "&H0000FF",   # Đỏ
            5: "&H0080FF",   # Cam
        }
        _pos_map = {0: "bottom", 1: "top", 2: "center"}
        _align_map = {0: "center", 1: "left", 2: "right"}
        _fx_map = {0: "none", 1: "fade", 2: "karaoke", 3: "word_pop",
                   4: "firework", 5: "popup", 6: "typewriter", 7: "bounce",
                   8: "slide_left", 9: "slide_right", 10: "glow"}

        # Thu thập config
        config = {
            "video": video,
            "output_dir": output_dir,
            "do_concat": do_concat,
            "group_size": self.spn_group.value(),
            "reencode": self.chk_reencode.isChecked(),
            "do_subtitle": do_sub,
            "model": self.cmb_model.currentText(),
            "language": self.cmb_lang.currentText(),
            "sub_font": self.cmb_sub_font.currentText(),
            "sub_size": int(self.spn_sub_size.value() * 1.33),  # Scale pixel→ASS point
            "sub_color": self.btn_text_color.current_color,
            "sub_bg_color": self.btn_bg_color.current_color,
            "kara_run_color": self.btn_kara_run_color.current_color,
            "kara_bg_color": self.btn_kara_bg_color.current_color,
            "sub_style": self.cmb_sub_style.currentIndex(),
            "sub_bold": getattr(self, 'btn_bold', type('o', (), {'isChecked': lambda: True})).isChecked(),
            "sub_italic": getattr(self, 'btn_italic', type('o', (), {'isChecked': lambda: False})).isChecked(),
            "sub_underline": getattr(self, 'btn_underline', type('o', (), {'isChecked': lambda: False})).isChecked(),
            "sub_border_w": getattr(self, 'spn_border_w', type('o', (), {'value': lambda: 3})).value(),
            "sub_position": _pos_map.get(self.cmb_sub_pos.currentIndex(), "bottom"),
            "sub_align": _align_map.get(self.cmb_sub_align.currentIndex(), "center"),
            "sub_fx": _fx_map.get(self.cmb_sub_fx.currentIndex(), "none"),
            "sub_margin_v": self.spn_margin_v.value(),
            "sub_x_offset": getattr(self, "_current_x_offset", 0),
            "sub_y_vi": self._calc_y_vi(),
            "aspect_idx": self.cmb_aspect.currentIndex(),
            "do_music": do_music,
            "music_path": self.txt_music.text(),
            "music_volume": self.spn_volume.value(),
            "music_start": self.spn_m_start.value(),
            "music_end": self.spn_m_end.value() if self.spn_m_end.value() > 0 else None,
            "fade_in": self.spn_fade_in.value(),
            "fade_out": self.spn_fade_out.value(),
            "batch_excel": self.txt_excel.text() if self.chk_excel_enabled.isChecked() else None,
            "col_prompt": self.txt_col_prompt.text(),
            "col_upload": self.txt_col_upload.text(),
            "start_id": int(self.txt_start_id.text() if self.txt_start_id.text().isdigit() else 1),
        }

        def worker():
            excel_wb = None
            excel_ws = None
            try:
                import shutil
                video_path = Path(config["video"])
                video_dir = video_path
                out_dir = Path(config["output_dir"])
                out_dir.mkdir(parents=True, exist_ok=True)
                tmp_dir = out_dir / "_tmp"
                if tmp_dir.exists():
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                tmp_dir.mkdir(exist_ok=True)

                # Xóa file cũ trong output
                for old in out_dir.glob("final_*.mp4"):
                    old.unlink(missing_ok=True)

                # ── XÁC ĐỊNH VIDEO ĐẦU VÀO (NORMAL vs EXCEL BATCH) ──
                working_groups = []
                cup = -1
                c_used = -1
                
                if config.get("batch_excel"):
                    import openpyxl, re
                    e_path = Path(config["batch_excel"])
                    sid = config.get("start_id", 1)
                    if e_path.exists():
                        self._log(f"📊 Đọc Excel Batch: {e_path.name} (ID từ: {sid})")
                        self._close_excel_if_open(str(e_path))
                        excel_wb = openpyxl.load_workbook(e_path)
                        excel_ws = excel_wb.active
                        headers = [str(c.value).strip().lower() for c in next(excel_ws.rows)]
                        cpr = headers.index(config["col_prompt"].lower()) if config["col_prompt"].lower() in headers else -1
                        cup = headers.index(config["col_upload"].lower()) if config["col_upload"].lower() in headers else -1
                        
                        # Tìm hoặc tạo cột video_used
                        col_used_name = "video_used"
                        if col_used_name in headers:
                            c_used = headers.index(col_used_name)
                        else:
                            # Tạo cột mới ở cuối
                            c_used = len(headers)
                            excel_ws.cell(row=1, column=c_used+1).value = "video_used"
                            headers.append(col_used_name)
                            self._log(f"📝 Tạo cột '{col_used_name}' mới")
                        
                        # Tìm cột video name
                        col_vname_name = "video name"
                        c_vname = headers.index(col_vname_name) if col_vname_name in headers else -1
                        
                        if cpr != -1 and cup != -1:
                            # Quét toàn bộ video_used để tìm ID lớn nhất đã dùng
                            max_used_id = sid - 1
                            if c_used != -1:
                                for scan_row in excel_ws.iter_rows(min_row=2):
                                    uv = scan_row[c_used].value if c_used < len(scan_row) else None
                                    if uv and str(uv).strip():
                                        ids_in_row = [int(x.strip()) for x in str(uv).split(",") if x.strip().isdigit()]
                                        if ids_in_row:
                                            max_used_id = max(max_used_id, max(ids_in_row))
                            
                            current_id = max(sid, max_used_id + 1)
                            self._log(f"📌 Video ID tiếp theo: {current_id}")
                            
                            for ridx, row in enumerate(excel_ws.iter_rows(min_row=2), start=2):
                                prompt = row[cpr].value
                                v_upload = row[cup].value

                                if not prompt or (v_upload and str(v_upload).strip()):
                                    continue
                                
                                nsc = len([line for line in str(prompt).splitlines() if line.strip()])
                                if nsc <= 0: continue
                                
                                # Đọc video_used để biết đã dùng video nào trước đó
                                used_val = excel_ws.cell(row=ridx, column=c_used+1).value
                                already_used = []
                                if used_val and str(used_val).strip():
                                    already_used = [int(x.strip()) for x in str(used_val).split(",") if x.strip().isdigit()]
                                
                                # Bỏ qua video đã dùng, lấy tiếp từ sau video cuối cùng
                                if already_used:
                                    start_from = max(already_used) + 1
                                    remaining = nsc - len(already_used)
                                    if remaining <= 0:
                                        continue  # Đã đủ video cho hàng này
                                    row_ids = list(range(start_from, start_from + remaining))
                                else:
                                    row_ids = list(range(current_id, current_id + nsc))
                                
                                sfiles = []
                                missing = False
                                for vid in row_ids:
                                    fnd = list(video_dir.glob(f"{vid}_1_*.mp4"))
                                    if not fnd: fnd = list(video_dir.glob(f"{vid}_1.mp4"))
                                    
                                    if fnd: sfiles.append(fnd[0])
                                    else: missing = True; break
                                
                                if missing:
                                    self._log(f"⚠️ Hàng {ridx}: Thiếu Video ID trong dãy {row_ids} → Bỏ qua")
                                    if not already_used:
                                        current_id += nsc
                                    continue
                                
                                # Lấy video name từ cột video name
                                vname = ""
                                if c_vname != -1:
                                    raw_vname = excel_ws.cell(row=ridx, column=c_vname+1).value
                                    if raw_vname:
                                        vname = str(raw_vname).split("#")[0].strip()
                                
                                # Tên file: {row-1}_{video_name}
                                row_num = ridx - 1  # hàng thứ bao nhiêu trừ 1
                                if vname:
                                    # Làm sạch tên file
                                    safe_name = re.sub(r'[<>:"/\\|?*]', '_', vname).strip()
                                    base_name = f"{row_num}_{safe_name}"
                                else:
                                    base_name = f"{row_num}"
                                
                                # Hàng hợp lệ - lưu thêm row_ids và already_used
                                working_groups.append((sfiles, ridx, base_name, nsc, row_ids, already_used))
                                if not already_used:
                                    current_id += nsc
                            self._log(f"→ Tìm được {len(working_groups)} hàng Excel mới")
                        else:
                            self._log(f"❌ Không tìm thấy cột")

                if not working_groups:
                    all_f = sorted(video_path.glob("*.mp4")) if video_path.is_dir() else [video_path]
                    if not config["do_concat"]:
                        for f in all_f: working_groups.append(([f], -1, f.stem, 1, [], []))
                    else:
                        gs = config["group_size"]
                        for i in range(0, len(all_f), gs):
                            g = all_f[i:i+gs]
                            working_groups.append((g, -1, f"merged_{i//gs+1:03d}", len(g), [], []))

                # ── BẮT ĐẦU XỬ LÝ TỪNG NHÓM ──
                for gi, (files, row_idx, base_name, nsc, row_ids, already_used) in enumerate(working_groups):
                    if self._check_stop(): break
                    self._log(f"{'━'*40}\n🚀 [{gi+1}/{len(working_groups)}] {base_name} | Cảnh: {nsc} | Video: {len(files)}")
                    
                    cur_vid = None
                    # NỐI
                    if len(files) > 1 or config["do_concat"]:
                        from video_concat import VideoConcat
                        vc = VideoConcat(log_callback=self._log)
                        out_p = tmp_dir / f"{base_name}_merged.mp4"
                        vc.concat_single_group(files, str(out_p), re_encode=config["reencode"])
                        if out_p.exists(): cur_vid = out_p
                        else: 
                            self._log("❌ Nối thất bại"); continue
                    else:
                        cur_vid = tmp_dir / files[0].name
                        shutil.copy2(files[0], cur_vid)
                    
                    current_file = cur_vid

                    # ══════════════════════════════════════════
                    # XỬ LÝ TỪNG VIDEO: Sub → Nhạc → Move
                    # ══════════════════════════════════════════
                    # ── PHỤ ĐỀ ──
                    if config["do_subtitle"] and current_file and not self._check_stop():
                        try:
                            sub_payload = {
                                "video": str(current_file),
                                "model": config["model"], "language": config["language"],
                                "burn": True, "mode": "single",
                                "font": config["sub_font"], "font_size": config["sub_size"],
                                "font_color": config["sub_color"], "bg_color": config.get("sub_bg_color", "#000000"),
                                "style_idx": config.get("sub_style", 0),
                                "is_bold": config.get("sub_bold", True),
                                "is_italic": config.get("sub_italic", False),
                                "is_underline": config.get("sub_underline", False),
                                "border_w": config.get("sub_border_w", 3),
                                "position": config["sub_position"],
                                "align": config["sub_align"], "fx": config["sub_fx"],
                                "margin_v": config["sub_margin_v"],
                                "x_offset": config.get("sub_x_offset", 0),
                                "y_vi": config.get("sub_y_vi"),
                                "auto_y": config.get("sub_y_vi") is None,
                                "aspect_idx": config.get("aspect_idx", 0),
                            }
                            sub_config = json.dumps(sub_payload, ensure_ascii=False)
                            worker_script = str(PROJECT_ROOT / "subtitle_worker.py")
                            proc = sp.Popen([sys.executable, worker_script, sub_config],
                                stdout=sp.PIPE, stderr=sp.PIPE, cwd=str(PROJECT_ROOT),
                                encoding="utf-8", errors="replace", creationflags=0x08000000)
                            self._sub_proc = proc
                            for line in proc.stdout:
                                line = line.strip()
                                if not line: continue
                                try:
                                    msg = json.loads(line)
                                    if msg.get("type") == "log": self._log_raw(msg["msg"])
                                    elif msg.get("type") == "error": self._log(f"❌ {msg['msg']}")
                                except json.JSONDecodeError: self._log_raw(line)
                            proc.wait()
                            self._sub_proc = None
                            burned = current_file.parent / f"{current_file.stem}_sub.mp4"
                            if burned.exists(): current_file = burned
                        except Exception as e: self._log(f"❌ Phụ đề lỗi: {e}")

                    # ── NHẠC NỀN ──
                    if config["do_music"] and current_file and not self._check_stop():
                        music = config["music_path"]
                        if music and Path(music).exists():
                            try:
                                from background_music import BackgroundMusic
                                bgm = BackgroundMusic(log_callback=self._log)
                                out_f = tmp_dir / f"bgm_{gi+1:03d}.mp4"
                                bgm.add_music(video_path=current_file, music_path=music, output_path=out_f,
                                    music_volume=config["music_volume"], music_start=config["music_start"],
                                    music_end=config["music_end"], fade_in=config["fade_in"], fade_out=config["fade_out"])
                                if out_f.exists(): current_file = out_f
                            except Exception as e: self._log(f"❌ Nhạc nền lỗi: {e}")

                    # MOVE TO OUTPUT
                    if current_file and current_file.exists() and not self._check_stop():
                        # Đặt tên file theo base_name (ID_TênVideo)
                        final_path = out_dir / f"{base_name}.mp4"
                        shutil.move(str(current_file), str(final_path))
                        self._log(f"✅ → {final_path.name}")
                        
                        # Ghi vào cột video_upload + video_used
                        if excel_ws and row_idx != -1 and cup != -1:
                            excel_ws.cell(row=row_idx, column=cup+1).value = str(final_path)
                            # Ghi danh sách video đã dùng (gộp cũ + mới)
                            if c_used != -1 and row_ids:
                                all_used = already_used + row_ids
                                excel_ws.cell(row=row_idx, column=c_used+1).value = ", ".join(str(x) for x in all_used)
                            self._close_excel_if_open(config["batch_excel"])
                            excel_wb.save(config["batch_excel"])

                # Dọn dẹp
                if tmp_dir.exists(): shutil.rmtree(tmp_dir, ignore_errors=True)
                if self._check_stop(): self._log("⏹ Đã dừng.")
                else: self._log(f"\\n✨ HOÀN THÀNH TẤT CẢ! → {out_dir}")

            except Exception as e:
                import traceback; traceback.print_exc()
                self._log(f"❌ Lỗi: {e}")
            finally:
                if 'excel_wb' in locals() and excel_wb: excel_wb.close()
                self._sig_done.emit("")

        threading.Thread(target=worker, daemon=True).start()




    def _check_stop(self):
        return not self._running

    def _stop_all(self):
        self._running = False
        if self._sub_proc:
            try:
                self._sub_proc.kill()
            except Exception:
                pass
        self._reset_run_btn()
        self._log("⏹ Đang dừng...")

    def _on_done(self, msg):
        self._running = False
        self._reset_run_btn()
        self._refresh_output_list()

    def _reset_run_btn(self):
        """Reset nút chạy về trạng thái ban đầu."""
        self.btn_run.setText("🚀 Chạy")
        self.btn_run.setObjectName("PrimaryButton")
        self.btn_run.style().unpolish(self.btn_run)
        self.btn_run.style().polish(self.btn_run)


# ════════════════════════════════════════════════════════════
# MAIN WINDOW
# ════════════════════════════════════════════════════════════
class AppMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Auto Video Editor")
        self.setMinimumSize(1100, 750)
        self.resize(1350, 880)
        self._build_ui()
        self._apply_style()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Header
        header = QWidget()
        header.setObjectName("AppHeader")
        header.setFixedHeight(52)
        hl = QHBoxLayout(header)
        hl.setContentsMargins(20, 0, 20, 0)
        title = QLabel("Auto Video Editor")
        title.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        title.setObjectName("MainTitle")
        hl.addWidget(title)
        hl.addStretch()
        main_layout.addWidget(header)

        # Tabs
        self.tabs = QTabWidget()
        self.tabs.setObjectName("MainTabs")
        self.tabs.setDocumentMode(True)

        # Tab 1: APP tiếng anh
        self.english_tab = TextOverlayEditor()
        self.tabs.addTab(self.english_tab, "🇬🇧 APP tiếng anh")

        # Tab 2: Video Tools (gộp 3 chức năng)
        self.tools_tab = VideoToolsTab()
        self.tabs.addTab(self.tools_tab, "🎬 Video Tools")

        main_layout.addWidget(self.tabs, 1)

    def _apply_style(self):
        self.setStyleSheet(APP_QSS)


# ════════════════════════════════════════════════════════════
# QSS
# ════════════════════════════════════════════════════════════
APP_QSS = """
QMainWindow { background: #0e0e1a; }

#AppHeader {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
        stop:0 #151530, stop:1 #0e0e22);
    border-bottom: 1px solid rgba(100,100,200,0.15);
}
#MainTitle { color: #e6e6ff; }

#MainTabs { background: #0e0e1a; }
#MainTabs::pane { background: #0e0e1a; border: none; }

#MainTabs QTabBar::tab {
    background: rgba(20,20,45,0.9);
    color: #8888bb;
    padding: 10px 22px;
    margin-right: 2px;
    border: 1px solid rgba(80,80,160,0.12);
    border-bottom: none;
    border-top-left-radius: 8px;
    border-top-right-radius: 8px;
    font-size: 12px;
    font-weight: 600;
    min-width: 120px;
}
#MainTabs QTabBar::tab:selected {
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #1a1a40, stop:1 #0e0e1a);
    color: #e0e0ff;
    border: 1px solid rgba(108,92,231,0.35);
    border-bottom: 2px solid #6c5ce7;
}
#MainTabs QTabBar::tab:hover:!selected {
    background: rgba(30,30,60,0.95);
    color: #b0b0dd;
}

#SettingsGroup {
    background: rgba(22,22,48,0.85);
    border: 1px solid rgba(90,90,180,0.15);
    border-radius: 10px;
    padding: 12px 12px 8px 12px;
    margin-top: 4px;
    font-weight: 600;
    color: #c0c0e8;
    font-size: 13px;
}

#FieldLabel { color: #9090c0; font-size: 12px; }

#TextInput {
    background: rgba(18,18,40,0.9);
    border: 1px solid rgba(90,90,170,0.2);
    border-radius: 6px;
    padding: 6px 10px;
    color: #d0d0ee;
    font-size: 12px;
    min-height: 20px;
}
#TextInput:focus { border: 1px solid rgba(108,92,231,0.5); }

#FontCombo, QComboBox {
    background: rgba(18,18,40,0.9);
    border: 1px solid rgba(90,90,170,0.2);
    border-radius: 6px;
    padding: 5px 10px;
    color: #d0d0ee;
    font-size: 12px;
    min-height: 22px;
}
QComboBox::drop-down {
    border: none; width: 24px;
    background: rgba(108,92,231,0.15);
    border-top-right-radius: 6px;
    border-bottom-right-radius: 6px;
}
QComboBox QAbstractItemView {
    background: #161630;
    border: 1px solid rgba(90,90,170,0.3);
    color: #d0d0ee;
    selection-background-color: rgba(108,92,231,0.4);
    font-size: 12px;
    padding: 4px;
}

#ValueSpin, QSpinBox, QDoubleSpinBox {
    background: rgba(18,18,40,0.9);
    border: 1px solid rgba(90,90,170,0.2);
    border-radius: 6px;
    padding: 4px 8px;
    color: #d0d0ee;
    font-size: 12px;
    min-height: 20px;
}

/* Feature checkbox (lớn, rõ) */
#FeatureCheck {
    color: #d0d0ff;
    font-size: 15px;
    font-weight: 700;
    spacing: 10px;
}
#FeatureCheck::indicator {
    width: 26px; height: 26px;
    border-radius: 6px;
    border: 2px solid rgba(108,92,231,0.5);
    background: rgba(18,18,40,0.9);
}
#FeatureCheck::indicator:checked {
    background: rgba(108,92,231,0.7);
    border: 2px solid #6c5ce7;
}

#OptionCheck, QCheckBox {
    color: #b0b0d8;
    font-size: 12px;
    spacing: 8px;
}
QCheckBox::indicator {
    width: 20px; height: 20px;
    border-radius: 5px;
    border: 2px solid rgba(90,90,170,0.3);
    background: rgba(18,18,40,0.9);
}
QCheckBox::indicator:checked {
    background: rgba(108,92,231,0.6);
    border: 2px solid #6c5ce7;
}

#PrimaryButton {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
        stop:0 #6c5ce7, stop:1 #5a4bd1);
    color: white;
    border: none;
    border-radius: 8px;
    padding: 10px 24px;
    font-weight: 600;
    font-size: 14px;
    min-height: 28px;
}
#PrimaryButton:hover {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
        stop:0 #7d6ff0, stop:1 #6c5ce7);
}
#PrimaryButton:pressed { background: #4a3cb8; }
#PrimaryButton:disabled { background: rgba(90,90,130,0.4); color: #666; }

#SecondaryButton {
    background: rgba(40,40,80,0.7);
    color: #b0b0d8;
    border: 1px solid rgba(90,90,170,0.2);
    border-radius: 6px;
    padding: 6px 14px;
    font-size: 12px;
    min-height: 22px;
}
#SecondaryButton:hover {
    background: rgba(60,60,110,0.8);
    border: 1px solid rgba(108,92,231,0.4);
    color: #d0d0ee;
}

#DangerButton {
    background: rgba(200,50,50,0.15);
    color: #ff6b6b;
    border: 1px solid rgba(200,50,50,0.3);
    border-radius: 6px;
    padding: 6px 14px;
    font-size: 12px;
    min-height: 22px;
}
#DangerButton:hover { background: rgba(200,50,50,0.3); }
#DangerButton:disabled {
    background: rgba(80,40,40,0.2);
    color: #665555;
    border: 1px solid rgba(100,50,50,0.15);
}

#LogPanel {
    background: rgba(10,10,25,0.95);
    border: 1px solid rgba(90,90,170,0.12);
    border-radius: 8px;
    color: #a0a0d0;
    padding: 8px;
    font-size: 10px;
}

#FileList {
    background: rgba(14,14,35,0.95);
    border: 1px solid rgba(90,90,170,0.15);
    border-radius: 8px;
    color: #c0c0e8;
    padding: 4px;
    font-size: 11px;
}
#FileList::item {
    padding: 4px 8px;
    border-radius: 4px;
}
#FileList::item:selected {
    background: rgba(108,92,231,0.4);
    color: #e8e8ff;
}
#FileList::item:hover {
    background: rgba(108,92,231,0.2);
}

#PreviewPanel { background: #0e0e1a; }

QSplitter::handle { background: rgba(90,90,170,0.1); width: 3px; }

/* Thanh tua video */
#SeekSlider::groove:horizontal {
    background: rgba(40,40,80,0.8);
    height: 6px;
    border-radius: 3px;
}
#SeekSlider::handle:horizontal {
    background: #6c5ce7;
    width: 14px;
    height: 14px;
    margin: -5px 0;
    border-radius: 7px;
    border: 2px solid rgba(108,92,231,0.8);
}
#SeekSlider::handle:horizontal:hover {
    background: #7d6ff0;
    width: 16px;
    height: 16px;
    margin: -6px 0;
    border-radius: 8px;
}
#SeekSlider::sub-page:horizontal {
    background: rgba(108,92,231,0.6);
    border-radius: 3px;
}

QScrollBar:vertical { background: transparent; width: 8px; }
QScrollBar::handle:vertical {
    background: rgba(108,92,231,0.3);
    border-radius: 4px; min-height: 30px;
}
QScrollBar::handle:vertical:hover { background: rgba(108,92,231,0.5); }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: transparent; }

QSpinBox::up-button, QDoubleSpinBox::up-button {
    background: rgba(108,92,231,0.2);
    border-left: 1px solid rgba(90,90,170,0.12);
    border-top-right-radius: 5px; width: 16px;
}
QSpinBox::up-button:hover, QDoubleSpinBox::up-button:hover { background: rgba(108,92,231,0.4); }
QSpinBox::down-button, QDoubleSpinBox::down-button {
    background: rgba(108,92,231,0.2);
    border-left: 1px solid rgba(90,90,170,0.12);
    border-bottom-right-radius: 5px; width: 16px;
}
QSpinBox::down-button:hover, QDoubleSpinBox::down-button:hover { background: rgba(108,92,231,0.4); }
QSpinBox::up-arrow, QDoubleSpinBox::up-arrow {
    image: none; width: 0; height: 0;
    border-left: 3px solid transparent;
    border-right: 3px solid transparent;
    border-bottom: 4px solid #9999cc;
}
QSpinBox::down-arrow, QDoubleSpinBox::down-arrow {
    image: none; width: 0; height: 0;
    border-left: 3px solid transparent;
    border-right: 3px solid transparent;
    border-top: 4px solid #9999cc;
}
"""


# ════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════
def run_app():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))

    if FONTS_DIR.exists():
        for f in FONTS_DIR.glob("*.ttf"):
            QFontDatabase.addApplicationFont(str(f))

    win = AppMainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    run_app()