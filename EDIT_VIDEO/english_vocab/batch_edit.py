# -*- coding: utf-8 -*-
"""
batch_edit.py
=============
Đọc Excel → ghép 4 video/chủ đề + chèn từ vựng → xuất file cuối.

Cách dùng:
  .venv/Scripts/python.exe batch_edit.py
"""

import sys, subprocess, shutil, re
from pathlib import Path

# Fix console encoding cho Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ════════════════════════════════════════════════════════════
# CẤU HÌNH
# ════════════════════════════════════════════════════════════
EXCEL_PATH  = Path(r"D:\AI\TOOL_VEO3_NEW\APP Tieng anh\output\100_topics_english_scripts_V1.xlsx")
_VIDEO_DIR1 = Path(r"D:\AI\Xay Kenh\DATA\100 Video\ENGLISH_ List 100 chủ đề\video")
_VIDEO_DIR2 = Path(r"D:\AI\Xay Kenh\DATA\100 Video\ENGLISH_ List 100 chu de\video")
VIDEO_DIR   = _VIDEO_DIR1 if _VIDEO_DIR1.exists() else _VIDEO_DIR2
OUTPUT_DIR  = Path(r"D:\AI\TOOL_VEO3_NEW\TEST_AUTO_edit\output\batch")

SCENES_PER_TOPIC = 4     # mỗi chủ đề gồm 4 video
SWITCH_TIME      = 4.0   # giây: chuyển từ vựng 1 → 2
W1_START = 0.0
W1_END   = 4.0
W2_START = 4.0
W2_END   = 99.0

# ── Style chữ ──
FONT_BOLD   = r"C\:/Windows/Fonts/arialbd.ttf"
FONT_ITALIC = r"C\:/Windows/Fonts/ariali.ttf"
SIZE_VI     = 62
SIZE_EN     = 62
SIZE_IPA    = 50
COLOR_TEXT  = "0xFFD700"
COLOR_BORDER= "0x000000"
BORDER_W    = 4
Y_VI        = 988
Y_EN        = 1058
Y_IPA       = 1130


# ════════════════════════════════════════════════════════════
# BƯỚC 1 — ĐỌC EXCEL
# ════════════════════════════════════════════════════════════
def read_excel(path: Path) -> list[dict]:
    """
    Đọc danh sách chủ đề từ Excel.
    Trả về list[{stt, topic, vocab: [{vi,en,ipa},...] }]

    Cấu trúc Excel:
      Col A = STT
      Col B = Chủ đề
      Col C = Từ vựng (mỗi word = 3 dòng: VI / EN / IPA)
    """
    try:
        import openpyxl
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    topics = []
    current = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        stt   = row[0] if len(row) > 0 else None
        topic = row[1] if len(row) > 1 else None
        vocab_raw = row[2] if len(row) > 2 else None

        # Nếu có STT → hàng đầu của chủ đề mới
        if stt is not None:
            if current:
                topics.append(current)
            current = {"stt": int(stt), "topic": str(topic).strip(), "vocab": []}

        # Parse vocab từ cell (mỗi word = 3 dòng: VI / EN / IPA)
        if vocab_raw and current is not None:
            lines = [l.strip() for l in str(vocab_raw).splitlines() if l.strip()]
            i = 0
            while i + 2 < len(lines):
                current["vocab"].append({
                    "vi":  lines[i],
                    "en":  lines[i+1],
                    "ipa": lines[i+2],
                })
                i += 3

    if current:
        topics.append(current)

    return topics


# ════════════════════════════════════════════════════════════
# BƯỚC 2 — TÌM VIDEO FILES CHO MỖI TOPIC
# ════════════════════════════════════════════════════════════
def find_videos_for_topic(stt: int, video_dir: Path) -> list[Path | None]:
    """
    Topic STT=N → cần video số (N-1)*4+1 đến N*4
    File tên dạng: {num}_1_{...}.mp4
    Trả về list 4 phần tử: Path nếu tìm thấy, None nếu thiếu.
    """
    base = (stt - 1) * SCENES_PER_TOPIC
    result = []
    for i in range(1, SCENES_PER_TOPIC + 1):
        num = base + i
        # Tìm file bắt đầu bằng "{num}_"
        matches = sorted(video_dir.glob(f"{num}_*.mp4"))
        result.append(matches[0] if matches else None)
    return result


# ════════════════════════════════════════════════════════════
# BƯỚC 3 — FFmpeg helpers
# ════════════════════════════════════════════════════════════
def get_ffmpeg() -> str:
    """Lấy đường dẫn ffmpeg — ưu tiên local folder, fallback system."""
    local_ff = Path(__file__).parent.parent / "ffmpeg" / "ffmpeg.exe"
    if local_ff.exists():
        return str(local_ff)
    ff = shutil.which("ffmpeg")
    if ff:
        return ff
    raise RuntimeError("Không tìm thấy ffmpeg! Hãy đặt ffmpeg.exe vào folder ffmpeg/")


def esc(s: str) -> str:
    s = s.replace("\\", "\\\\")
    s = s.replace("'",  "\u2019")
    s = s.replace(":",  "\\:")
    return s


def build_drawtext(vi_text, en_text, ipa_text, t_start, t_end) -> str:
    enable = f"between(t\\,{t_start}\\,{t_end})"

    def dt(text, font, size, y):
        return (
            f"drawtext=fontfile='{font}'"
            f":text='{esc(text)}'"
            f":fontsize={size}"
            f":fontcolor={COLOR_TEXT}"
            f":bordercolor={COLOR_BORDER}"
            f":borderw={BORDER_W}"
            f":x=(w-text_w)/2"
            f":y={y}"
            f":enable='{enable}'"
        )

    return ",".join([
        dt(vi_text,  FONT_BOLD,   SIZE_VI,  Y_VI),
        dt(en_text,  FONT_BOLD,   SIZE_EN,  Y_EN),
        dt(ipa_text, FONT_ITALIC, SIZE_IPA, Y_IPA),
    ])


def process_scene(ffmpeg: str, src: Path, vocab_pair: list[dict], out: Path):
    """Chèn 2 từ vựng lên 1 video scene rồi lưu ra out."""
    if not vocab_pair:
        shutil.copy2(src, out)
        return

    parts = []
    if len(vocab_pair) >= 1:
        v = vocab_pair[0]
        parts.append(build_drawtext(v["vi"], v["en"], v["ipa"], W1_START, W1_END))
    if len(vocab_pair) >= 2:
        v = vocab_pair[1]
        parts.append(build_drawtext(v["vi"], v["en"], v["ipa"], W2_START, W2_END))

    vf = ",".join(parts) + ",format=yuv420p"

    cmd = [
        ffmpeg, "-y", "-i", str(src),
        "-vf", vf,
        "-c:v", "libx264", "-preset", "veryfast", "-crf", "18",
        "-pix_fmt", "yuv420p",
        "-color_range", "tv", "-colorspace", "bt709",
        "-color_primaries", "bt709", "-color_trc", "bt709",
        "-c:a", "copy",
        str(out),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    if r.returncode != 0:
        raise RuntimeError(f"FFmpeg error:\n{r.stderr[-600:]}")


def concat_videos(ffmpeg: str, parts: list[Path], output: Path):
    list_file = output.parent / "_concat_list.txt"
    with open(list_file, "w", encoding="utf-8") as f:
        for p in parts:
            f.write(f"file '{str(p).replace(chr(92), '/')}'\n")
    cmd = [
        ffmpeg, "-y", "-f", "concat", "-safe", "0",
        "-i", str(list_file), "-c", "copy", str(output),
    ]
    r = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace")
    list_file.unlink(missing_ok=True)
    if r.returncode != 0:
        raise RuntimeError(f"Concat error:\n{r.stderr[-400:]}")


# ════════════════════════════════════════════════════════════
# BƯỚC 4 — GHI STATUS VÀO EXCEL
# ════════════════════════════════════════════════════════════
def write_status_excel(path: Path, status_map: dict[int, str]):
    """Ghi cột Status vào Excel theo STT → status string."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    # Tìm hoặc tạo header "Status"
    status_col = None
    for col in range(1, ws.max_column + 2):
        v = ws.cell(1, col).value
        if v and "status" in str(v).lower():
            status_col = col
            break
    if status_col is None:
        status_col = ws.max_column + 1
        ws.cell(1, status_col, "Status")

    # Ghi status theo STT
    for row in ws.iter_rows(min_row=2):
        stt_val = row[0].value
        if stt_val is not None:
            stt = int(stt_val)
            if stt in status_map:
                ws.cell(row[0].row, status_col, status_map[stt])

    wb.save(str(path))


# ════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════
def main():
    print("=" * 65)
    print("  BATCH VIDEO EDITOR — 100 Chủ đề")
    print("=" * 65)

    if not EXCEL_PATH.exists():
        print(f"❌ Không tìm thấy Excel: {EXCEL_PATH}")
        sys.exit(1)
    if not VIDEO_DIR.exists():
        print(f"❌ Không tìm thấy video folder: {VIDEO_DIR}")
        sys.exit(1)

    # ── Đọc Excel ──
    print(f"\n📊 Đọc Excel: {EXCEL_PATH.name} ...")
    topics = read_excel(EXCEL_PATH)
    print(f"   ✅ {len(topics)} chủ đề")

    ffmpeg = get_ffmpeg()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    tmp_dir = OUTPUT_DIR / "_tmp"
    tmp_dir.mkdir(exist_ok=True)

    status_map = {}   # stt → status string
    ok_count   = 0
    skip_count = 0

    for topic in topics:
        stt      = topic["stt"]
        name     = topic["topic"]
        vocab    = topic["vocab"]   # 6 từ vựng
        safe_name = re.sub(r'[\\/:*?"<>|]', '_', name)

        print(f"\n{'─'*60}")
        print(f"[{stt:3d}] {name}")

        # ── Tìm 4 video ──
        videos = find_videos_for_topic(stt, VIDEO_DIR)
        missing = [i+1 for i, v in enumerate(videos) if v is None]

        if missing:
            msg = f"Bỏ qua — thiếu video: scene {missing}"
            print(f"   ⚠️  {msg}")
            status_map[stt] = msg
            skip_count += 1
            continue

        print(f"   🎬 Video: {[v.name for v in videos]}")
        print(f"   📝 Vocab: {[w['vi'] for w in vocab]}")

        # Chia vocab: scene 1→từ0,1 | scene 2→từ2,3 | scene 3→từ4,5 | scene 4→[]
        vocab_per_scene = [
            vocab[0:2],
            vocab[2:4],
            vocab[4:6],
            [],          # scene 4 = CTA
        ]

        # ── Xử lý từng scene ──
        processed_parts = []
        try:
            for si, (vpath, vpair) in enumerate(zip(videos, vocab_per_scene)):
                tmp_out = tmp_dir / f"stt{stt}_scene{si+1}.mp4"
                process_scene(ffmpeg, vpath, vpair, tmp_out)
                processed_parts.append(tmp_out)
                print(f"   ✅ Scene {si+1} xong")

            # ── Ghép 4 scene thành 1 video ──
            final_name = f"{stt}_{safe_name}.mp4"
            final_path = OUTPUT_DIR / final_name
            concat_videos(ffmpeg, processed_parts, final_path)

            # Dọn temp scene files
            for p in processed_parts:
                p.unlink(missing_ok=True)

            print(f"   🎉 Hoàn thành → {final_name}")
            status_map[stt] = f"OK → {final_name}"
            ok_count += 1

        except Exception as e:
            print(f"   ❌ Lỗi: {e}")
            status_map[stt] = f"Lỗi: {str(e)[:100]}"
            skip_count += 1
            for p in processed_parts:
                p.unlink(missing_ok=True)

    # ── Dọn tmp ──
    shutil.rmtree(tmp_dir, ignore_errors=True)

    # ── Ghi status vào Excel ──
    print(f"\n📝 Ghi status vào Excel ...")
    write_status_excel(EXCEL_PATH, status_map)

    # ── Tổng kết ──
    print(f"\n{'='*65}")
    print(f"  ✅ Thành công : {ok_count} chủ đề")
    print(f"  ⚠️  Bỏ qua    : {skip_count} chủ đề")
    print(f"  📁 Output    : {OUTPUT_DIR}")
    print(f"{'='*65}\n")


if __name__ == "__main__":
    main()
