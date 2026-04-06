# -*- coding: utf-8 -*-
"""
text_overlay_editor.py
======================
Giao diện chỉnh sửa text overlay cho video từ vựng.

Cấu trúc đơn giản:
  ┌─────────────────────────────────────────────┐
  │  STYLE CHUNG (font, màu, viền, cỡ chữ)     │
  │  TỪ 1: VI / EN / IPA + thời gian           │
  │  TỪ 2: VI / EN / IPA + thời gian           │
  │  VỊ TRÍ (Y cho 3 dòng VI, EN, IPA)         │
  │  NÚT: Lưu / Copy FFmpeg                    │
  ├─────────────────────────────────────────────┤
  │  PREVIEW CANVAS (kéo thả text)              │
  └─────────────────────────────────────────────┘

Cách dùng:
  .venv/Scripts/python.exe text_overlay_editor.py
"""

import sys
import json
import os
import re
import shutil
import subprocess
import threading
from pathlib import Path

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QComboBox, QSpinBox, QDoubleSpinBox,
    QColorDialog, QFileDialog, QGroupBox, QScrollArea,
    QSplitter, QFrame, QLineEdit, QSizePolicy, QMessageBox,
    QToolButton, QTextEdit
)
from PyQt6.QtCore import (
    Qt, QPoint, QPointF, QRect, QRectF, QSize, QTimer, pyqtSignal
)
from PyQt6.QtGui import (
    QFont, QColor, QPainter, QPen, QBrush, QFontDatabase,
    QPixmap, QImage, QLinearGradient, QPainterPath, QFontMetrics, QCursor
)

PROJECT_ROOT = Path(__file__).parent.parent   # TEST_AUTO_edit/
CONFIG_FILE = Path(__file__).parent / "text_overlay_config.json"
FONTS_DIR = PROJECT_ROOT / "fonts"


# ════════════════════════════════════════════════════════════
# PREVIEW CANVAS — Kéo thả text trên preview
# ════════════════════════════════════════════════════════════
class PreviewCanvas(QWidget):
    """Canvas xem trước text overlay, hỗ trợ kéo thả 3 dòng text."""

    position_changed = pyqtSignal(int)  # new Y value

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumSize(400, 500)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.setCursor(QCursor(Qt.CursorShape.OpenHandCursor))

        # Canvas = kích thước video thật (mặc định 720x1280)
        self.canvas_w = 720
        self.canvas_h = 1280

        # Dữ liệu hiển thị
        self.display_data = None  # dict từ MainWindow

        # Drag state
        self._dragging = False
        self._drag_offset_y = 0

        # Background
        self._bg_pixmap = None

        self.setMouseTracking(True)

    def set_background(self, path: str):
        if path and os.path.exists(path):
            self._bg_pixmap = QPixmap(path)
        else:
            self._bg_pixmap = None
        self.update()

    def set_display_data(self, data: dict):
        self.display_data = data
        self.update()

    def _canvas_rect(self) -> QRect:
        w, h = self.width(), self.height()
        ratio = self.canvas_w / self.canvas_h
        wr = w / max(h, 1)
        if wr > ratio:
            ch = h
            cw = int(h * ratio)
        else:
            cw = w
            ch = int(w / ratio)
        return QRect((w - cw) // 2, (h - ch) // 2, cw, ch)

    def _real_to_canvas_y(self, ry):
        r = self._canvas_rect()
        return int(ry / self.canvas_h * r.height()) + r.y()

    def _canvas_to_real_y(self, cy):
        r = self._canvas_rect()
        return int((cy - r.y()) / max(r.height(), 1) * self.canvas_h)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        p.setRenderHint(QPainter.RenderHint.TextAntialiasing)

        rect = self._canvas_rect()

        # ── Nền ──
        if self._bg_pixmap:
            p.drawPixmap(rect, self._bg_pixmap)
        else:
            grad = QLinearGradient(QPointF(rect.x(), rect.y()),
                                   QPointF(rect.x() + rect.width(), rect.y() + rect.height()))
            grad.setColorAt(0, QColor(25, 25, 45))
            grad.setColorAt(0.5, QColor(18, 20, 38))
            grad.setColorAt(1, QColor(12, 12, 28))
            p.fillRect(rect, grad)

            # Lưới nhẹ
            pen = QPen(QColor(255, 255, 255, 12))
            p.setPen(pen)
            sx = max(rect.width() // 10, 1)
            for i in range(1, 10):
                x = rect.x() + i * sx
                p.drawLine(x, rect.y(), x, rect.y() + rect.height())
            sy = max(rect.height() // 18, 1)
            for i in range(1, 18):
                y = rect.y() + i * sy
                p.drawLine(rect.x(), y, rect.x() + rect.width(), y)

        # Viền canvas
        p.setPen(QPen(QColor(80, 80, 130, 100), 2))
        p.drawRect(rect)

        # ── Vẽ text ──
        if self.display_data:
            d = self.display_data
            # Chọn từ nào hiển thị (hiện tại hiện từ 1 trên preview)
            word = d.get("word1", {})
            if not word.get("vi"):
                word = d.get("word2", {})

            vi = word.get("vi", "Tiếng Việt")
            en = word.get("en", "English")
            ipa = word.get("ipa", "/ipa/")

            self._draw_three_lines(p, rect, vi, en, ipa, d)

        # Info góc
        p.setPen(QColor(150, 150, 180, 160))
        p.setFont(QFont("Segoe UI", 8))
        p.drawText(rect.x() + 6, rect.y() + 14, f"{self.canvas_w}×{self.canvas_h}")

        p.end()

    def _draw_three_lines(self, p: QPainter, rect: QRect, vi, en, ipa, d):
        """Vẽ 3 dòng text (VI, EN, IPA) lên canvas."""
        font_name = d.get("font_name", "Arial")
        font_size = d.get("font_size", 62)
        ipa_size = max(font_size - 12, 20)  # giống batch_edit IPA nhỏ hơn
        bold = d.get("bold", True)
        italic_ipa = True
        text_color = QColor(d.get("text_color", "#FFD700"))
        border_color = QColor(d.get("border_color", "#000000"))
        border_w = d.get("border_width", 4)
        y_vi = d.get("y_vi", 988)
        y_en = d.get("y_en", 1058)
        y_ipa = d.get("y_ipa", 1130)

        scale = rect.width() / max(self.canvas_w, 1)

        lines_info = [
            (vi,  font_size, bold, False,      y_vi),
            (en,  font_size, bold, False,      y_en),
            (ipa, ipa_size,  False, italic_ipa, y_ipa),
        ]

        self._text_rects = []  # lưu vùng text cho drag

        for text, fsize, is_bold, is_italic, real_y in lines_info:
            if not text:
                self._text_rects.append(None)
                continue

            ss = max(1, int(fsize * scale))
            font = QFont("Arial", 1)  # dùng Arial giống FFmpeg
            font.setPixelSize(ss)     # pixel chính xác
            font.setBold(is_bold)
            font.setItalic(is_italic)

            fm = QFontMetrics(font)
            tw = fm.horizontalAdvance(text)
            th = fm.height()

            # Canh giữa X
            cx = rect.x() + (rect.width() - tw) // 2
            # FFmpeg y = top of text, QPainter y = baseline
            # → cộng ascent để khớp FFmpeg
            cy = self._real_to_canvas_y(real_y) + fm.ascent()

            # Lưu rect (dùng top-edge cho drag)
            self._text_rects.append(QRect(cx - 4, cy - fm.ascent(), tw + 8, th + 8))

            # Stroke
            sbw = max(1, int(border_w * scale))
            path = QPainterPath()
            path.addText(float(cx), float(cy), font, text)

            stroke_pen = QPen(border_color, sbw * 2)
            stroke_pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
            p.setPen(stroke_pen)
            p.setBrush(Qt.BrushStyle.NoBrush)
            p.drawPath(path)

            # Fill
            p.setPen(Qt.PenStyle.NoPen)
            p.setBrush(QBrush(text_color))
            p.drawPath(path)

        # Highlight vùng text khi drag
        if self._dragging:
            for tr in self._text_rects:
                if tr:
                    p.setPen(QPen(QColor(0, 180, 255, 120), 1, Qt.PenStyle.DashLine))
                    p.setBrush(Qt.BrushStyle.NoBrush)
                    p.drawRect(tr)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            # Cho phép drag nhóm 3 dòng text
            pos = event.pos()
            if hasattr(self, '_text_rects') and self._text_rects:
                for tr in self._text_rects:
                    if tr and tr.contains(pos):
                        self._dragging = True
                        self._drag_offset_y = pos.y() - tr.center().y()
                        self.setCursor(QCursor(Qt.CursorShape.ClosedHandCursor))
                        return

    def mouseMoveEvent(self, event):
        if self._dragging and self.display_data:
            pos = event.pos()
            # Tính delta Y thực tế
            d = self.display_data
            old_y_vi = d.get("y_vi", 988)
            old_y_en = d.get("y_en", 1058)
            old_y_ipa = d.get("y_ipa", 1130)

            # Sử dụng dòng đầu tiên (VI) làm anchor
            if hasattr(self, '_text_rects') and self._text_rects and self._text_rects[0]:
                old_canvas_y = self._text_rects[0].center().y()
                new_canvas_y = pos.y() - self._drag_offset_y
                delta_canvas = new_canvas_y - old_canvas_y
                delta_real = self._canvas_to_real_y(
                    self._canvas_rect().y() + delta_canvas
                ) - self._canvas_to_real_y(self._canvas_rect().y())

                new_y_vi = max(0, min(1900, old_y_vi + delta_real))
                gap_en = old_y_en - old_y_vi
                gap_ipa = old_y_ipa - old_y_vi
                new_y_en = new_y_vi + gap_en
                new_y_ipa = new_y_vi + gap_ipa

                d["y_vi"] = new_y_vi
                d["y_en"] = new_y_en
                d["y_ipa"] = new_y_ipa

                self.position_changed.emit(new_y_vi)
                self.update()
        else:
            pos = event.pos()
            over = False
            if hasattr(self, '_text_rects'):
                for tr in self._text_rects:
                    if tr and tr.contains(pos):
                        over = True
                        break
            self.setCursor(QCursor(
                Qt.CursorShape.PointingHandCursor if over
                else Qt.CursorShape.OpenHandCursor
            ))

    def mouseReleaseEvent(self, event):
        if self._dragging:
            self._dragging = False
            self.setCursor(QCursor(Qt.CursorShape.OpenHandCursor))
            self.update()


# ════════════════════════════════════════════════════════════
# COLOR BUTTON
# ════════════════════════════════════════════════════════════
class ColorButton(QPushButton):
    color_changed = pyqtSignal(str)

    def __init__(self, color="#FFD700", parent=None):
        super().__init__(parent)
        self._color = QColor(color)
        self.setFixedSize(44, 32)
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.clicked.connect(self._pick)
        self._refresh()

    def _refresh(self):
        c = self._color.name()
        self.setStyleSheet(f"""
            QPushButton {{
                background: {c};
                border: 2px solid rgba(255,255,255,0.3);
                border-radius: 6px;
            }}
            QPushButton:hover {{ border: 2px solid rgba(255,255,255,0.7); }}
        """)

    def _pick(self):
        c = QColorDialog.getColor(self._color, self, "Chọn màu")
        if c.isValid():
            self._color = c
            self._refresh()
            self.color_changed.emit(c.name())

    def color(self) -> str:
        return self._color.name()

    def set_color(self, hex_c: str):
        self._color = QColor(hex_c)
        self._refresh()


# ════════════════════════════════════════════════════════════
# MAIN WINDOW
# ════════════════════════════════════════════════════════════
class TextOverlayEditor(QWidget):
    # Signal thread-safe: True = stopped, False = done
    _sig_batch_finished = pyqtSignal(bool)

    def __init__(self):
        super().__init__()
        self.setWindowTitle("✨ Text Overlay Editor")
        self.setMinimumSize(1100, 750)
        self.resize(1300, 850)
        self._sig_batch_finished.connect(self._on_batch_finished)

        self._build_ui()
        self._apply_style()
        self._load_config()
        self._refresh_preview()

    def _build_ui(self):
        main_layout = QHBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.setHandleWidth(3)

        # ══════════════════════════════════════
        # LEFT — SIDEBAR
        # ══════════════════════════════════════
        sidebar = QWidget()
        sidebar.setObjectName("Sidebar")
        sidebar.setMinimumWidth(480)
        sidebar.setMaximumWidth(700)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_layout.setSpacing(0)

        # ── Header ──
        header = QWidget()
        header.setObjectName("SidebarHeader")
        hl = QVBoxLayout(header)
        hl.setContentsMargins(18, 16, 18, 12)
        title = QLabel("✨ Text Overlay Editor")
        title.setObjectName("AppTitle")
        title.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
        hl.addWidget(title)
        sub = QLabel("Chỉnh sửa text overlay cho video từ vựng")
        sub.setObjectName("AppSubtitle")
        hl.addWidget(sub)
        sidebar_layout.addWidget(header)

        # ── Scroll area ──
        scroll = QScrollArea()
        scroll.setObjectName("ContentScroll")
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        content = QWidget()
        content.setObjectName("ContentWidget")
        cl = QVBoxLayout(content)
        cl.setContentsMargins(14, 10, 14, 10)
        cl.setSpacing(12)

        LBL_W = 48  # chiều rộng label cố định cho thẳng hàng
        SPN_W = 65  # chiều rộng spinbox nhỏ gọn

        def lbl_fixed(text, w=LBL_W):
            l = QLabel(text)
            l.setObjectName("FieldLabel")
            l.setFixedWidth(w)
            return l

        def small_spin(lo, hi, val, suffix=""):
            s = self._spin(lo, hi, val, suffix)
            s.setFixedWidth(SPN_W)
            return s

        def small_spin_na(lo, hi, val, suffix=""):
            s = self._spin_no_auto(lo, hi, val, suffix)
            s.setFixedWidth(SPN_W)
            return s

        def small_dspin(lo, hi, val, suffix=""):
            s = self._dspin(lo, hi, val, suffix)
            s.setFixedWidth(SPN_W)
            return s

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # NHÓM 0: BATCH PROCESSING
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        grp_batch = QGroupBox("📁 Dữ liệu & Batch")
        grp_batch.setObjectName("SettingsGroup")
        gb = QVBoxLayout(grp_batch)
        gb.setSpacing(5)

        # Excel
        rb1 = QHBoxLayout()
        rb1.addWidget(lbl_fixed("Excel:"))
        self.txt_excel = QLineEdit()
        self.txt_excel.setObjectName("TextInput")
        self.txt_excel.setPlaceholderText("Chọn file Excel...")
        self.txt_excel.setReadOnly(True)
        rb1.addWidget(self.txt_excel, 1)
        btn_excel = QPushButton("📂 Chọn")
        btn_excel.setObjectName("SecondaryButton")
        btn_excel.setFixedWidth(60)
        btn_excel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_excel.clicked.connect(self._browse_excel)
        rb1.addWidget(btn_excel)
        gb.addLayout(rb1)

        # Video folder
        rb2 = QHBoxLayout()
        rb2.addWidget(lbl_fixed("Video:"))
        self.txt_video_dir = QLineEdit()
        self.txt_video_dir.setObjectName("TextInput")
        self.txt_video_dir.setPlaceholderText("Chọn folder video...")
        self.txt_video_dir.setReadOnly(True)
        rb2.addWidget(self.txt_video_dir, 1)
        btn_vdir = QPushButton("📂 Chọn")
        btn_vdir.setObjectName("SecondaryButton")
        btn_vdir.setFixedWidth(60)
        btn_vdir.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_vdir.clicked.connect(self._browse_video_dir)
        rb2.addWidget(btn_vdir)
        gb.addLayout(rb2)

        # Output folder
        rb_out = QHBoxLayout()
        rb_out.addWidget(lbl_fixed("Output:"))
        self.txt_output_dir = QLineEdit()
        self.txt_output_dir.setObjectName("TextInput")
        self.txt_output_dir.setPlaceholderText("Chọn folder lưu output...")
        self.txt_output_dir.setReadOnly(True)
        rb_out.addWidget(self.txt_output_dir, 1)
        btn_odir = QPushButton("📂 Chọn")
        btn_odir.setObjectName("SecondaryButton")
        btn_odir.setFixedWidth(60)
        btn_odir.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_odir.clicked.connect(self._browse_output_dir)
        rb_out.addWidget(btn_odir)
        gb.addLayout(rb_out)

        # Số scene + nút chạy
        rb3 = QHBoxLayout()
        rb3.addWidget(lbl_fixed("Scene:"))
        self.spn_scenes = QSpinBox()
        self.spn_scenes.setObjectName("ValueSpin")
        self.spn_scenes.setRange(1, 20)
        self.spn_scenes.setValue(4)
        self.spn_scenes.setFixedWidth(SPN_W)
        rb3.addWidget(self.spn_scenes)
        rb3.addStretch()
        self.btn_run_batch = QPushButton("🚀 Chạy Batch")
        self.btn_run_batch.setObjectName("RunBatchBtn")
        self.btn_run_batch.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_run_batch.clicked.connect(self._run_batch)
        rb3.addWidget(self.btn_run_batch)
        self.btn_stop_batch = QPushButton("⏹ Dừng")
        self.btn_stop_batch.setObjectName("DangerButton")
        self.btn_stop_batch.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_stop_batch.clicked.connect(self._stop_batch)
        self.btn_stop_batch.setEnabled(False)
        rb3.addWidget(self.btn_stop_batch)
        gb.addLayout(rb3)

        cl.addWidget(grp_batch)

        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        # NHÓM 1: CẤU HÌNH TEXT
        # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        grp_cfg = QGroupBox("🎨 Cấu hình Text")
        grp_cfg.setObjectName("SettingsGroup")
        gc = QVBoxLayout(grp_cfg)
        gc.setSpacing(5)

        # Row 1: Font + Size + Bold
        r1 = QHBoxLayout()
        r1.addWidget(lbl_fixed("Font:"))
        self.cmb_font = QComboBox()
        self.cmb_font.setObjectName("FontCombo")
        self._populate_fonts()
        self.cmb_font.currentTextChanged.connect(self._refresh_preview)
        r1.addWidget(self.cmb_font, 1)
        r1.addSpacing(4)
        r1.addWidget(self._lbl("Cỡ:"))
        self.spn_font_size = small_spin(10, 200, 62, " px")
        r1.addWidget(self.spn_font_size)
        self.btn_bold = QPushButton("B")
        self.btn_bold.setObjectName("StyleToggleBtn")
        self.btn_bold.setCheckable(True)
        self.btn_bold.setChecked(True)
        self.btn_bold.setFixedSize(28, 28)
        self.btn_bold.setFont(QFont("Arial", 10, QFont.Weight.Bold))
        self.btn_bold.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_bold.toggled.connect(self._refresh_preview)
        r1.addWidget(self.btn_bold)
        gc.addLayout(r1)

        # Row 2: Màu + Viền + Dày
        r2 = QHBoxLayout()
        r2.addWidget(lbl_fixed("Chữ:"))
        self.btn_text_color = ColorButton("#FFD700")
        self.btn_text_color.color_changed.connect(lambda _: self._refresh_preview())
        r2.addWidget(self.btn_text_color)
        r2.addSpacing(10)
        r2.addWidget(self._lbl("Viền:"))
        self.btn_border_color = ColorButton("#000000")
        self.btn_border_color.color_changed.connect(lambda _: self._refresh_preview())
        r2.addWidget(self.btn_border_color)
        r2.addSpacing(10)
        r2.addWidget(self._lbl("Dày:"))
        self.spn_border_w = small_spin(0, 20, 4, " px")
        r2.addWidget(self.spn_border_w)
        r2.addStretch()
        gc.addLayout(r2)

        # Row 3: Thời gian Từ 1
        rt1 = QHBoxLayout()
        rt1.addWidget(lbl_fixed("⏱ Từ 1:"))
        self.spn_start1 = small_dspin(0, 999, 0.0, " s")
        rt1.addWidget(self.spn_start1)
        rt1.addWidget(self._lbl("→"))
        self.spn_end1 = small_dspin(0, 999, 4.0, " s")
        rt1.addWidget(self.spn_end1)
        rt1.addStretch()
        gc.addLayout(rt1)

        # Row 4: Thời gian Từ 2
        rt2 = QHBoxLayout()
        rt2.addWidget(lbl_fixed("⏱ Từ 2:"))
        self.spn_start2 = small_dspin(0, 999, 4.0, " s")
        rt2.addWidget(self.spn_start2)
        rt2.addWidget(self._lbl("→"))
        self.spn_end2 = small_dspin(0, 999, 8.0, " s")
        rt2.addWidget(self.spn_end2)
        rt2.addStretch()
        gc.addLayout(rt2)

        # Row 5: Y positions + Gap
        r_pos = QHBoxLayout()
        r_pos.addWidget(lbl_fixed("Y-VI:"))
        self.spn_y_vi = small_spin_na(0, 1920, 988, "")
        self.spn_y_vi.valueChanged.connect(self._on_y_vi_changed)
        r_pos.addWidget(self.spn_y_vi)
        r_pos.addWidget(self._lbl("EN:"))
        self.spn_y_en = small_spin_na(0, 1920, 1058, "")
        self.spn_y_en.valueChanged.connect(self._refresh_preview)
        r_pos.addWidget(self.spn_y_en)
        r_pos.addWidget(self._lbl("IPA:"))
        self.spn_y_ipa = small_spin_na(0, 1920, 1130, "")
        self.spn_y_ipa.valueChanged.connect(self._refresh_preview)
        r_pos.addWidget(self.spn_y_ipa)
        r_pos.addSpacing(6)
        r_pos.addWidget(self._lbl("Gap:"))
        self.spn_line_gap = small_spin_na(20, 300, 70, "")
        self.spn_line_gap.valueChanged.connect(self._on_gap_changed)
        r_pos.addWidget(self.spn_line_gap)
        r_pos.addStretch()
        gc.addLayout(r_pos)

        # Row 6: Nút căn vị trí
        r_btn = QHBoxLayout()
        r_btn.addWidget(lbl_fixed(""))
        btn_center_v = QPushButton("⊞ Căn giữa dọc")
        btn_center_v.setObjectName("SecondaryButton")
        btn_center_v.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_center_v.setFixedHeight(26)
        btn_center_v.clicked.connect(self._center_vertical)
        r_btn.addWidget(btn_center_v)
        btn_bottom = QPushButton("⬇ Đặt dưới")
        btn_bottom.setObjectName("SecondaryButton")
        btn_bottom.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_bottom.setFixedHeight(26)
        btn_bottom.clicked.connect(self._place_bottom)
        r_btn.addWidget(btn_bottom)
        r_btn.addStretch()
        gc.addLayout(r_btn)

        cl.addWidget(grp_cfg)

        cl.addStretch()
        scroll.setWidget(content)
        sidebar_layout.addWidget(scroll, 1)

        # ── Bottom bar ──
        bottom = QWidget()
        bottom.setObjectName("BottomBar")
        bl = QHBoxLayout(bottom)
        bl.setContentsMargins(14, 10, 14, 12)
        bl.setSpacing(8)

        btn_bg = QPushButton("🖼️ Ảnh nền")
        btn_bg.setObjectName("SecondaryButton")
        btn_bg.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_bg.clicked.connect(self._choose_bg)
        bl.addWidget(btn_bg)

        btn_save = QPushButton("💾 Lưu")
        btn_save.setObjectName("PrimaryButton")
        btn_save.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_save.clicked.connect(self._save_config)
        bl.addWidget(btn_save)

        btn_export = QPushButton("📋 Copy FFmpeg")
        btn_export.setObjectName("PrimaryButton")
        btn_export.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_export.clicked.connect(self._export_ffmpeg)
        bl.addWidget(btn_export)

        btn_load = QPushButton("📂 Tải")
        btn_load.setObjectName("SecondaryButton")
        btn_load.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_load.clicked.connect(self._load_from_file)
        bl.addWidget(btn_load)

        sidebar_layout.addWidget(bottom)

        # ══════════════════════════════════════
        # RIGHT — PREVIEW + LOG
        # ══════════════════════════════════════
        right_panel = QWidget()
        right_panel.setObjectName("PreviewWrapper")
        rp = QVBoxLayout(right_panel)
        rp.setContentsMargins(10, 10, 10, 10)
        rp.setSpacing(6)

        pv_header = QLabel("👁️ Preview — Kéo thả text để thay đổi vị trí Y")
        pv_header.setObjectName("PreviewTitle")
        pv_header.setFont(QFont("Segoe UI", 10, QFont.Weight.DemiBold))
        rp.addWidget(pv_header)

        # Splitter dọc: Preview chiếm nhiều, Log ít
        right_splitter = QSplitter(Qt.Orientation.Vertical)

        self.canvas = PreviewCanvas()
        self.canvas.setMinimumHeight(300)
        self.canvas.position_changed.connect(self._on_canvas_drag)
        right_splitter.addWidget(self.canvas)

        # Info
        info = QWidget()
        info.setObjectName("InfoBar")
        il = QHBoxLayout(info)
        il.setContentsMargins(8, 4, 8, 4)
        self.lbl_info = QLabel("Y: 988 / 1058 / 1130")
        self.lbl_info.setObjectName("InfoLabel")
        il.addWidget(self.lbl_info)
        il.addStretch()

        # ── LOG PANEL ──
        log_widget = QWidget()
        log_layout = QVBoxLayout(log_widget)
        log_layout.setContentsMargins(0, 0, 0, 0)
        log_layout.setSpacing(4)
        log_header = QLabel("📋 Nhật ký hoạt động")
        log_header.setObjectName("PreviewTitle")
        log_header.setFont(QFont("Segoe UI", 10, QFont.Weight.DemiBold))
        log_layout.addWidget(log_header)

        self.txt_log = QTextEdit()
        self.txt_log.setObjectName("LogPanel")
        self.txt_log.setReadOnly(True)
        self.txt_log.setFont(QFont("Consolas", 9))
        self.txt_log.setMinimumHeight(100)
        log_layout.addWidget(self.txt_log, 1)
        right_splitter.addWidget(log_widget)

        # Tỉ lệ: preview 70%, log 30%
        right_splitter.setStretchFactor(0, 7)
        right_splitter.setStretchFactor(1, 3)

        rp.addWidget(info)
        rp.addWidget(right_splitter, 1)

        # ── Splitter ──
        splitter.addWidget(sidebar)
        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 1)
        splitter.setSizes([650, 450])
        main_layout.addWidget(splitter)

        self._show_word = 1

    # ── Helper tạo widget ──
    def _lbl(self, text):
        l = QLabel(text)
        l.setObjectName("FieldLabel")
        return l

    def _spin(self, mn, mx, val, suffix):
        s = QSpinBox()
        s.setObjectName("ValueSpin")
        s.setRange(mn, mx)
        s.setValue(val)
        if suffix:
            s.setSuffix(suffix)
        s.valueChanged.connect(self._refresh_preview)
        return s

    def _spin_no_auto(self, mn, mx, val, suffix):
        """SpinBox không tự connect _refresh_preview (sẽ connect riêng)."""
        s = QSpinBox()
        s.setObjectName("ValueSpin")
        s.setRange(mn, mx)
        s.setValue(val)
        if suffix:
            s.setSuffix(suffix)
        return s

    def _dspin(self, mn, mx, val, suffix):
        s = QDoubleSpinBox()
        s.setObjectName("ValueSpin")
        s.setRange(mn, mx)
        s.setValue(val)
        s.setSuffix(suffix)
        s.setDecimals(1)
        s.setSingleStep(0.5)
        s.valueChanged.connect(self._refresh_preview)
        return s

    def _populate_fonts(self):
        fonts_set = set()
        for family in QFontDatabase.families():
            fonts_set.add(family)
        if FONTS_DIR.exists():
            for f in FONTS_DIR.glob("*.ttf"):
                fid = QFontDatabase.addApplicationFont(str(f))
                if fid >= 0:
                    for fam in QFontDatabase.applicationFontFamilies(fid):
                        fonts_set.add(fam)

        priority = ["Arial", "Poppins", "Nunito", "Baloo 2", "Quicksand",
                     "Comfortaa", "Cabin", "Roboto", "Be Vietnam Pro",
                     "Segoe UI", "Tahoma"]
        result = []
        for p in priority:
            if p in fonts_set:
                result.append(p)
                fonts_set.discard(p)
        result.extend(sorted(fonts_set))
        self.cmb_font.addItems(result)

    # ── Khoảng cách dòng ──
    def _on_y_vi_changed(self, val):
        """Khi thay đổi Y_VI → tự tính Y_EN, Y_IPA theo khoảng cách dòng."""
        gap = self.spn_line_gap.value()
        self.spn_y_en.blockSignals(True)
        self.spn_y_ipa.blockSignals(True)
        self.spn_y_en.setValue(val + gap)
        self.spn_y_ipa.setValue(val + gap * 2)
        self.spn_y_en.blockSignals(False)
        self.spn_y_ipa.blockSignals(False)
        self._refresh_preview()

    def _on_gap_changed(self, gap):
        """Khi thay đổi khoảng cách dòng → tự tính lại Y_EN, Y_IPA."""
        y_vi = self.spn_y_vi.value()
        self.spn_y_en.blockSignals(True)
        self.spn_y_ipa.blockSignals(True)
        self.spn_y_en.setValue(y_vi + gap)
        self.spn_y_ipa.setValue(y_vi + gap * 2)
        self.spn_y_en.blockSignals(False)
        self.spn_y_ipa.blockSignals(False)
        self._refresh_preview()

    def _center_vertical(self):
        """Căn giữa nhóm 3 dòng text theo chiều dọc (canvas 1920px)."""
        gap = self.spn_line_gap.value()
        total_h = gap * 2  # khoảng cách từ dòng đầu đến dòng cuối
        y_vi = (1920 - total_h) // 2
        self.spn_y_vi.setValue(y_vi)  # sẽ trigger _on_y_vi_changed

    def _place_bottom(self):
        """Đặt nhóm 3 dòng text ở phần dưới video (khoảng 75% chiều cao)."""
        gap = self.spn_line_gap.value()
        y_vi = 1920 - gap * 2 - 150  # cách đáy ~150px
        self.spn_y_vi.setValue(max(0, y_vi))

    def _log(self, msg: str):
        """Ghi log vào panel nhật ký (thread-safe)."""
        from datetime import datetime
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        print(f"[LOG] {line}", flush=True)
        # Thread-safe: dùng QMetaObject invoke
        from PyQt6.QtCore import QMetaObject, Qt as QtNS, Q_ARG
        QMetaObject.invokeMethod(
            self.txt_log, "append",
            QtNS.ConnectionType.QueuedConnection,
            Q_ARG(str, line))

    # ── Collect config ──
    def _get_config(self) -> dict:
        return {
            "font_name": self.cmb_font.currentText(),
            "font_size": self.spn_font_size.value(),
            "bold": self.btn_bold.isChecked(),
            "text_color": self.btn_text_color.color(),
            "border_color": self.btn_border_color.color(),
            "border_width": self.spn_border_w.value(),
            "line_gap": self.spn_line_gap.value(),
            "y_vi": self.spn_y_vi.value(),
            "y_en": self.spn_y_en.value(),
            "y_ipa": self.spn_y_ipa.value(),
            "word1": {
                "time_start": self.spn_start1.value(),
                "time_end": self.spn_end1.value(),
            },
            "word2": {
                "time_start": self.spn_start2.value(),
                "time_end": self.spn_end2.value(),
            },
            "excel_path": self.txt_excel.text(),
            "video_dir": self.txt_video_dir.text(),
            "output_dir": self.txt_output_dir.text(),
            "scenes": self.spn_scenes.value(),
        }

    # Text mẫu cố định cho preview
    SAMPLE_VI = "Kem đánh răng"
    SAMPLE_EN = "Toothpaste"
    SAMPLE_IPA = "/ˈtuːθ.peɪst/"

    # ── Preview ──
    def _refresh_preview(self):
        cfg = self._get_config()
        display = dict(cfg)
        display["word1"] = {
            "vi": self.SAMPLE_VI,
            "en": self.SAMPLE_EN,
            "ipa": self.SAMPLE_IPA,
        }
        self.canvas.set_display_data(display)
        self.lbl_info.setText(
            f"Y: {cfg['y_vi']} / {cfg['y_en']} / {cfg['y_ipa']}")

    def _on_canvas_drag(self, new_y_vi):
        d = self.canvas.display_data
        if d:
            self.spn_y_vi.blockSignals(True)
            self.spn_y_en.blockSignals(True)
            self.spn_y_ipa.blockSignals(True)
            self.spn_y_vi.setValue(d["y_vi"])
            self.spn_y_en.setValue(d["y_en"])
            self.spn_y_ipa.setValue(d["y_ipa"])
            self.spn_y_vi.blockSignals(False)
            self.spn_y_en.blockSignals(False)
            self.spn_y_ipa.blockSignals(False)
            self.lbl_info.setText(
                f"Y: {d['y_vi']} / {d['y_en']} / {d['y_ipa']}")

    # ── Background ──
    def _choose_bg(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn ảnh nền (screenshot video)", "",
            "Images (*.png *.jpg *.jpeg *.bmp *.webp);;All (*)")
        if path:
            self.canvas.set_background(path)

    # ── Save / Load ──
    def _save_config(self):
        cfg = self._get_config()
        try:
            CONFIG_FILE.write_text(
                json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
            QMessageBox.information(self, "✅ Thành công",
                                    f"Đã lưu cấu hình!\n{CONFIG_FILE}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", str(e))

    def _load_config(self):
        if not CONFIG_FILE.exists():
            return
        try:
            cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            # Apply values
            idx = self.cmb_font.findText(cfg.get("font_name", "Arial"))
            if idx >= 0:
                self.cmb_font.setCurrentIndex(idx)
            self.spn_font_size.setValue(cfg.get("font_size", 62))
            self.btn_bold.setChecked(cfg.get("bold", True))
            self.btn_text_color.set_color(cfg.get("text_color", "#FFD700"))
            self.btn_border_color.set_color(cfg.get("border_color", "#000000"))
            self.spn_border_w.setValue(cfg.get("border_width", 4))
            self.spn_line_gap.setValue(cfg.get("line_gap", 70))
            self.spn_y_vi.setValue(cfg.get("y_vi", 988))
            self.spn_y_en.setValue(cfg.get("y_en", 1058))
            self.spn_y_ipa.setValue(cfg.get("y_ipa", 1130))

            w1 = cfg.get("word1", {})
            self.spn_start1.setValue(w1.get("time_start", 0.0))
            self.spn_end1.setValue(w1.get("time_end", 4.0))

            w2 = cfg.get("word2", {})
            self.spn_start2.setValue(w2.get("time_start", 4.0))
            self.spn_end2.setValue(w2.get("time_end", 8.0))

            # Paths
            if cfg.get("excel_path"):
                self.txt_excel.setText(cfg["excel_path"])
            if cfg.get("video_dir"):
                self.txt_video_dir.setText(cfg["video_dir"])
            if cfg.get("output_dir"):
                self.txt_output_dir.setText(cfg["output_dir"])
            if cfg.get("scenes"):
                self.spn_scenes.setValue(cfg["scenes"])

            # Detect resolution nếu video dir đã có
            if cfg.get("video_dir") and Path(cfg["video_dir"]).exists():
                self._detect_and_update_resolution(cfg["video_dir"])
        except Exception:
            pass

    def _load_from_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn file cấu hình", str(CONFIG_FILE.parent), "JSON (*.json)")
        if path:
            try:
                old = CONFIG_FILE
                import shutil
                shutil.copy2(path, CONFIG_FILE)
                self._load_config()
                self._refresh_preview()
                QMessageBox.information(self, "✅", "Đã tải cấu hình!")
            except Exception as e:
                QMessageBox.critical(self, "Lỗi", str(e))

    # ── Export FFmpeg ──
    def _export_ffmpeg(self):
        cfg = self._get_config()
        font_path = self._resolve_font(cfg)

        def esc(s):
            return s.replace("\\", "\\\\").replace("'", "\u2019").replace(":", "\\:")

        def dt(vi, en, ipa, ts, te):
            enable = f"between(t\\,{ts}\\,{te})"
            tc = cfg["text_color"].lstrip("#")
            bc = cfg["border_color"].lstrip("#")
            bw = cfg["border_width"]
            fs = cfg["font_size"]
            parts = []
            for text, size, y, is_italic in [
                (vi,  fs, cfg["y_vi"],  False),
                (en,  fs, cfg["y_en"],  False),
                (ipa, fs, cfg["y_ipa"], True),
            ]:
                fp = font_path if not is_italic else font_path.replace("bd.", "i.")
                parts.append(
                    f"drawtext=fontfile='{fp}'"
                    f":text='{esc(text)}'"
                    f":fontsize={size}"
                    f":fontcolor=0x{tc}"
                    f":bordercolor=0x{bc}"
                    f":borderw={bw}"
                    f":x=(w-text_w)/2"
                    f":y={y}"
                    f":enable='{enable}'"
                )
            return ",".join(parts)

        w1 = cfg["word1"]
        w2 = cfg["word2"]
        result_parts = []
        if w1["vi"]:
            result_parts.append(dt(w1["vi"], w1["en"], w1["ipa"],
                                    w1["time_start"], w1["time_end"]))
        if w2["vi"]:
            result_parts.append(dt(w2["vi"], w2["en"], w2["ipa"],
                                    w2["time_start"], w2["time_end"]))

        result = ",".join(result_parts)
        QApplication.clipboard().setText(result)
        QMessageBox.information(self, "📋 Đã copy!",
                                f"FFmpeg drawtext filter đã copy vào clipboard!\n\n{result[:200]}...")

    def _resolve_font(self, cfg):
        bold = cfg.get("bold", True)
        suffix = "bd" if bold else ""
        return f"C\\:/Windows/Fonts/arial{suffix}.ttf"

    # ── Browse paths ──
    def _auto_save(self):
        """Tự động lưu config (không popup)."""
        try:
            cfg = self._get_config()
            CONFIG_FILE.write_text(
                json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass

    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Chọn file Excel", "", "Excel (*.xlsx *.xls);;All (*)")
        if path:
            self.txt_excel.setText(path)
            self._auto_save()

    def _browse_video_dir(self):
        path = QFileDialog.getExistingDirectory(self, "Chọn folder video")
        if path:
            self.txt_video_dir.setText(path)
            self._detect_and_update_resolution(path)
            self._auto_save()

    def _browse_output_dir(self):
        path = QFileDialog.getExistingDirectory(self, "Chọn folder lưu video output")
        if path:
            self.txt_output_dir.setText(path)
            self._auto_save()

    # ══════════════════════════════════════════════════════════
    # DETECT VIDEO RESOLUTION
    # ══════════════════════════════════════════════════════════
    def _detect_and_update_resolution(self, video_dir_path: str):
        """Detect WxH từ video đầu tiên, cập nhật canvas + spinbox."""
        vdir = Path(video_dir_path)
        if not vdir.exists():
            return
        # Tìm 1 file mp4 bất kỳ
        mp4s = sorted(vdir.glob("*.mp4"))
        if not mp4s:
            mp4s = sorted(vdir.rglob("*.mp4"))
        if not mp4s:
            return
        src = str(mp4s[0])
        try:
            ffmpeg = self._get_ffmpeg()
        except Exception:
            return
        try:
            r = subprocess.run(
                [ffmpeg, "-i", src],
                capture_output=True, text=True,
                encoding="utf-8", errors="replace")
            m = re.search(r'(\d{3,4})x(\d{3,4})', r.stderr)
            if m:
                w, h = int(m.group(1)), int(m.group(2))
                old_h = self.canvas.canvas_h
                self.canvas.canvas_w = w
                self.canvas.canvas_h = h
                # Cập nhật max cho spinbox Y
                self.spn_y_vi.setMaximum(h)
                self.spn_y_en.setMaximum(h)
                self.spn_y_ipa.setMaximum(h)
                # Scale Y nếu resolution thay đổi
                cur_vi = self.spn_y_vi.value()
                cur_en = self.spn_y_en.value()
                cur_ipa = self.spn_y_ipa.value()
                gap = self.spn_line_gap.value()
                # Nếu Y bị clamp (tất cả = h hoặc cao hơn 90% h) → đặt lại
                if cur_vi >= h * 0.9 or cur_en >= h * 0.9 or cur_ipa >= h * 0.9:
                    # Đặt text ở 60% chiều cao
                    base_y = int(h * 0.6)
                    self.spn_y_vi.setValue(base_y)
                    self.spn_y_en.setValue(base_y + gap)
                    self.spn_y_ipa.setValue(base_y + gap * 2)
                    self._log(f"📐 Video: {w}×{h} → auto Y={base_y}/{base_y+gap}/{base_y+gap*2}")
                elif old_h != h and old_h > 0:
                    ratio = h / old_h
                    self.spn_y_vi.setValue(int(cur_vi * ratio))
                    self.spn_y_en.setValue(int(cur_en * ratio))
                    self.spn_y_ipa.setValue(int(cur_ipa * ratio))
                    self._log(f"📐 Video: {w}×{h} (scale Y ×{ratio:.2f})")
                else:
                    self._log(f"📐 Video: {w}×{h}")
                self.canvas.update()
                self._refresh_preview()
        except Exception:
            pass

    # ══════════════════════════════════════════════════════════
    # BATCH PROCESSING
    # ══════════════════════════════════════════════════════════
    def _get_ffmpeg(self) -> str:
        # 1. Folder local trong dự án
        local_ff = PROJECT_ROOT / "ffmpeg" / "ffmpeg.exe"
        if local_ff.exists():
            return str(local_ff)
        # 2. PATH hệ thống
        ff = shutil.which("ffmpeg")
        if ff:
            return ff
        raise RuntimeError("Không tìm thấy ffmpeg! Hãy đặt ffmpeg.exe vào folder ffmpeg/")

    def _read_excel(self, path: Path) -> list:
        try:
            import openpyxl
        except ImportError:
            subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
            import openpyxl
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        topics = []
        current = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            stt = row[0] if len(row) > 0 else None
            topic = row[1] if len(row) > 1 else None
            vocab_raw = row[2] if len(row) > 2 else None
            if stt is not None:
                if current:
                    topics.append(current)
                current = {"stt": int(stt), "topic": str(topic).strip(), "vocab": []}
            if vocab_raw and current is not None:
                lines = [l.strip() for l in str(vocab_raw).splitlines() if l.strip()]
                i = 0
                while i + 2 < len(lines):
                    current["vocab"].append({
                        "vi": lines[i], "en": lines[i+1], "ipa": lines[i+2]
                    })
                    i += 3
        if current:
            topics.append(current)
        return topics

    def _find_videos(self, stt: int, video_dir: Path, scenes: int) -> list:
        base = (stt - 1) * scenes
        result = []
        for i in range(1, scenes + 1):
            num = base + i
            matches = sorted(video_dir.glob(f"{num}_*.mp4"))
            result.append(matches[0] if matches else None)
        return result

    def _esc(self, s: str) -> str:
        s = s.replace("\\", "\\\\")
        s = s.replace("'", "\u2019")
        s = s.replace(":", "\\:")
        return s

    def _build_drawtext(self, vi, en, ipa, t_start, t_end, cfg):
        enable = "between(t" + chr(92) + "," + str(t_start) + chr(92) + "," + str(t_end) + ")"
        font_path = self._resolve_font(cfg)
        font_italic = font_path.replace("bd.", "i.") if "bd." in font_path else font_path
        fs = cfg["font_size"]
        tc = cfg["text_color"].lstrip("#")
        bc = cfg["border_color"].lstrip("#")
        bw = cfg["border_width"]

        def dt(text, fp, y):
            return (
                f"drawtext=fontfile='{fp}'"
                f":text='{self._esc(text)}'"
                f":fontsize={fs}"
                f":fontcolor=0x{tc}"
                f":bordercolor=0x{bc}"
                f":borderw={bw}"
                f":x=(w-text_w)/2"
                f":y={y}"
                f":enable='{enable}'"
            )
        return ",".join([
            dt(vi, font_path, cfg["y_vi"]),
            dt(en, font_path, cfg["y_en"]),
            dt(ipa, font_italic, cfg["y_ipa"]),
        ])

    def _process_scene(self, ffmpeg, src, vocab_pair, out, cfg):
        if not vocab_pair:
            shutil.copy2(src, out)
            return
        parts = []
        if len(vocab_pair) >= 1:
            v = vocab_pair[0]
            parts.append(self._build_drawtext(
                v["vi"], v["en"], v["ipa"],
                cfg["word1"]["time_start"], cfg["word1"]["time_end"], cfg))
        if len(vocab_pair) >= 2:
            v = vocab_pair[1]
            parts.append(self._build_drawtext(
                v["vi"], v["en"], v["ipa"],
                cfg["word2"]["time_start"], cfg["word2"]["time_end"], cfg))
        vf = ",".join(parts) + ",format=yuv420p"
        cmd = [
            ffmpeg, "-y", "-i", str(src), "-vf", vf,
            "-c:v", "libx264", "-preset", "veryfast", "-crf", "18",
            "-pix_fmt", "yuv420p", "-color_range", "tv",
            "-colorspace", "bt709", "-color_primaries", "bt709",
            "-color_trc", "bt709", "-c:a", "copy", str(out),
        ]
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        self._batch_process = proc
        stdout, stderr = proc.communicate()
        self._batch_process = None
        if proc.returncode != 0:
            raise RuntimeError(f"FFmpeg error:\n{stderr.decode('utf-8', errors='replace')[-500:]}")

    def _concat_videos(self, ffmpeg, parts, output):
        list_file = output.parent / "_concat_list.txt"
        with open(list_file, "w", encoding="utf-8") as f:
            for p in parts:
                f.write(f"file '{str(p).replace(chr(92), '/')}'\n")
        cmd = [ffmpeg, "-y", "-f", "concat", "-safe", "0",
               "-i", str(list_file), "-c", "copy", str(output)]
        r = subprocess.run(cmd, capture_output=True, text=True,
                           encoding="utf-8", errors="replace")
        list_file.unlink(missing_ok=True)
        if r.returncode != 0:
            raise RuntimeError(f"Concat error:\n{r.stderr[-400:]}")

    def _write_status_excel(self, path, status_map):
        import openpyxl
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        status_col = None
        for col in range(1, ws.max_column + 2):
            v = ws.cell(1, col).value
            if v and "status" in str(v).lower():
                status_col = col
                break
        if status_col is None:
            status_col = ws.max_column + 1
            ws.cell(1, status_col, "Status")
        for row in ws.iter_rows(min_row=2):
            stt_val = row[0].value
            if stt_val is not None and int(stt_val) in status_map:
                ws.cell(row[0].row, status_col, status_map[int(stt_val)])
        wb.save(str(path))

    def _run_batch(self):
        excel_path = self.txt_excel.text().strip()
        video_dir = self.txt_video_dir.text().strip()
        output_path = self.txt_output_dir.text().strip()
        if not excel_path or not Path(excel_path).exists():
            QMessageBox.warning(self, "Thiếu dữ liệu", "Chọn file Excel trước!")
            return
        if not video_dir or not Path(video_dir).exists():
            QMessageBox.warning(self, "Thiếu dữ liệu", "Chọn folder video trước!")
            return
        if not output_path:
            output_path = str(Path(excel_path).parent / "output_batch")

        self._batch_stop = False
        self._batch_process = None
        self.btn_run_batch.setEnabled(False)
        self.btn_run_batch.setText("⏳ Đang xử lý...")
        self.btn_stop_batch.setEnabled(True)
        self._log("🚀 Bắt đầu batch processing...")

        cfg = self._get_config()
        scenes = self.spn_scenes.value()

        def worker():
            try:
                ffmpeg = self._get_ffmpeg()
                self._log(f"🔧 FFmpeg: {ffmpeg}")
            except Exception as e:
                import traceback; traceback.print_exc()
                self._log(f"❌ FFmpeg không tìm thấy: {e}")
                self._sig_batch_finished.emit(False)
                return
            try:
                # Truyền config UI vào batch_edit module
                from english_vocab import batch_edit as be
                font_bold = self._resolve_font(cfg)
                be.FONT_BOLD = font_bold
                be.FONT_ITALIC = font_bold.replace("bd.", "i.") if "bd." in font_bold else font_bold
                be.SIZE_VI = cfg["font_size"]
                be.SIZE_EN = cfg["font_size"]
                be.SIZE_IPA = max(cfg["font_size"] - 12, 20)
                be.COLOR_TEXT = "0x" + cfg["text_color"].lstrip("#")
                be.COLOR_BORDER = "0x" + cfg["border_color"].lstrip("#")
                be.BORDER_W = cfg["border_width"]
                be.Y_VI = cfg["y_vi"]
                be.Y_EN = cfg["y_en"]
                be.Y_IPA = cfg["y_ipa"]
                be.SWITCH_TIME = cfg["word1"]["time_end"]
                be.W1_START = cfg["word1"]["time_start"]
                be.W1_END = cfg["word1"]["time_end"]
                be.W2_START = cfg["word2"]["time_start"]
                be.W2_END = cfg["word2"]["time_end"]
                self._log(f"⚙️ Config: Y={be.Y_VI}/{be.Y_EN}/{be.Y_IPA} Size={be.SIZE_VI} Time={be.W1_START}-{be.W1_END}/{be.W2_START}-{be.W2_END}")

                topics = self._read_excel(Path(excel_path))
                self._log(f"📊 Đọc Excel: {len(topics)} chủ đề")

                # Đọc status cũ từ Excel → bỏ qua topic đã done
                existing_status = self._read_excel_status(Path(excel_path))
                self._log(f"📋 Đã done: {sum(1 for v in existing_status.values() if 'OK' in str(v))} chủ đề")

                output_dir = Path(output_path)
                output_dir.mkdir(parents=True, exist_ok=True)
                self._log(f"📁 Output: {output_dir}")
                tmp_dir = output_dir / "_tmp"
                tmp_dir.mkdir(exist_ok=True)

                ok_count = 0
                skip_count = 0

                for ti, topic in enumerate(topics):
                    if self._batch_stop:
                        self._log("⏹ Đã dừng bởi người dùng!")
                        break

                    stt = topic["stt"]
                    name = topic["topic"]
                    vocab = topic["vocab"]
                    safe_name = re.sub(r'[\\/:*?"<>|]', '_', name)

                    # Bỏ qua topic đã done
                    old_status = existing_status.get(stt, "")
                    if "OK" in str(old_status):
                        self._log(f"── [{stt}] {name} → ⏭ đã xong, bỏ qua")
                        ok_count += 1
                        continue

                    self._log(f"── [{stt}] {name}")

                    videos = self._find_videos(stt, Path(video_dir), scenes)
                    missing = [i+1 for i, v in enumerate(videos) if v is None]

                    if missing:
                        msg = f"Thiếu scene: {missing}"
                        self._log(f"   ⚠️ {msg}")
                        self._write_single_status(Path(excel_path), stt, msg)
                        skip_count += 1
                        continue

                    try:
                        processed_parts = []
                        for si in range(scenes):
                            if self._batch_stop:
                                break
                            start_idx = si * 2
                            vpair = vocab[start_idx:start_idx+2]
                            scene_out = tmp_dir / f"stt{stt}_s{si}.mp4"
                            be.process_scene(ffmpeg, videos[si], vpair, scene_out)
                            processed_parts.append(scene_out)

                        if self._batch_stop:
                            self._log("⏹ Đã dừng bởi người dùng!")
                            # Cleanup partial
                            for p in processed_parts:
                                p.unlink(missing_ok=True)
                            break

                        # Concat các scene đã chèn text
                        final = output_dir / f"{stt}_{safe_name}.mp4"
                        if len(processed_parts) == 1:
                            shutil.copy2(processed_parts[0], final)
                        else:
                            self._concat_videos(ffmpeg, processed_parts, final)

                        # Cleanup
                        for p in processed_parts:
                            p.unlink(missing_ok=True)

                        self._log(f"   ✅ Xong → {final.name}")
                        # Ghi status ngay vào Excel
                        self._write_single_status(Path(excel_path), stt, f"OK → {final.name}")
                        ok_count += 1

                    except Exception as e:
                        if self._batch_stop:
                            self._log("⏹ Đã dừng bởi người dùng!")
                            break
                        self._log(f"   ❌ Lỗi: {e}")
                        self._write_single_status(Path(excel_path), stt, f"Lỗi: {str(e)[:80]}")
                        skip_count += 1

                shutil.rmtree(tmp_dir, ignore_errors=True)
                self._log(f"✅ Hoàn thành! OK: {ok_count}, Bỏ qua: {skip_count}")
                self._log(f"📁 Output: {output_dir}")

                if not self._batch_stop:
                    QTimer.singleShot(0, lambda: QMessageBox.information(
                        self, "✅ Hoàn thành",
                        f"Thành công: {ok_count}\nBỏ qua: {skip_count}\n"
                        f"Output: {output_dir}"))
            except Exception as e:
                import traceback; traceback.print_exc()
                self._log(f"❌ Lỗi nghiêm trọng: {e}")
            finally:
                self._sig_batch_finished.emit(self._batch_stop)

        threading.Thread(target=worker, daemon=True).start()

    def _on_batch_finished(self, was_stopped: bool):
        """Slot chạy trên main thread khi batch kết thúc."""
        self.btn_stop_batch.setEnabled(False)
        self.btn_run_batch.setEnabled(True)
        if was_stopped:
            self.btn_run_batch.setText("▶ Tiếp tục")
        else:
            self.btn_run_batch.setText("🚀 Chạy Batch")

    def _stop_batch(self):
        self._batch_stop = True
        self._log("⏹ Đang dừng... chờ lệnh hiện tại hoàn thành.")
        if self._batch_process:
            try:
                self._batch_process.kill()
            except Exception:
                pass
        self.btn_stop_batch.setEnabled(False)

    def _read_excel_status(self, path: Path) -> dict:
        """Đọc cột Status từ Excel → {stt: status_string}"""
        import openpyxl
        result = {}
        try:
            wb = openpyxl.load_workbook(str(path))
            ws = wb.active
            status_col = None
            for col in range(1, ws.max_column + 1):
                v = ws.cell(1, col).value
                if v and "status" in str(v).lower():
                    status_col = col
                    break
            if status_col:
                for row in range(2, ws.max_row + 1):
                    stt = ws.cell(row, 1).value
                    status = ws.cell(row, status_col).value
                    if stt is not None and status:
                        result[int(stt)] = str(status)
            wb.close()
        except Exception:
            pass
        return result

    def _write_single_status(self, path: Path, stt: int, status: str):
        """Ghi status 1 topic vào Excel ngay lập tức."""
        import openpyxl
        try:
            wb = openpyxl.load_workbook(str(path))
            ws = wb.active
            status_col = None
            for col in range(1, ws.max_column + 2):
                v = ws.cell(1, col).value
                if v and "status" in str(v).lower():
                    status_col = col
                    break
            if status_col is None:
                status_col = ws.max_column + 1
                ws.cell(1, status_col, "Status")
            for row in range(2, ws.max_row + 1):
                cell_stt = ws.cell(row, 1).value
                if cell_stt is not None and int(cell_stt) == stt:
                    ws.cell(row, status_col, status)
                    break
            wb.save(str(path))
            wb.close()
        except Exception as e:
            print(f"[WARN] Không ghi được status: {e}", flush=True)

    def _get_duration(self, ffmpeg, path) -> float:
        """Lấy duration video bằng ffmpeg -i (không cần ffprobe)."""
        # Thử ffprobe trước
        ffprobe = ffmpeg.replace("ffmpeg", "ffprobe")
        if Path(ffprobe).exists():
            cmd = [ffprobe, "-v", "error", "-show_entries",
                   "format=duration", "-of", "csv=p=0", str(path)]
            r = subprocess.run(cmd, capture_output=True, text=True,
                               encoding="utf-8", errors="replace")
            try:
                return float(r.stdout.strip())
            except ValueError:
                pass
        # Fallback: dùng ffmpeg -i
        cmd = [ffmpeg, "-i", str(path)]
        r = subprocess.run(cmd, capture_output=True, text=True,
                           encoding="utf-8", errors="replace")
        # Parse "Duration: 00:00:08.23" từ stderr
        m = re.search(r"Duration:\s*(\d+):(\d+):(\d+)\.(\d+)", r.stderr)
        if m:
            h, mi, s, cs = int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))
            return h * 3600 + mi * 60 + s + cs / 100.0
        return 8.0  # fallback

    # ── QSS ──
    def _apply_style(self):
        self.setStyleSheet(QSS)


# ════════════════════════════════════════════════════════════
# QSS THEME
# ════════════════════════════════════════════════════════════
QSS = """
TextOverlayEditor { background: #0e0e1a; }

#Sidebar {
    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
        stop:0 #151528, stop:1 #0c0c1c);
    border-right: 1px solid rgba(90,90,170,0.18);
}
#SidebarHeader {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
        stop:0 #191935, stop:1 #111128);
    border-bottom: 1px solid rgba(90,90,180,0.15);
}
#AppTitle { color: #e6e6ff; }
#AppSubtitle { color: #8585b5; font-size: 10px; }

/* Groups */
#SettingsGroup {
    background: rgba(22,22,48,0.85);
    border: 1px solid rgba(90,90,180,0.15);
    border-radius: 10px;
    padding: 10px 8px 6px 8px;
    margin-top: 6px;
    font-weight: 600;
    color: #c0c0e8;
    font-size: 12px;
}
#SettingsGroup::title {
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 6px;
    color: #d0d0f0;
}

#FieldLabel {
    color: #9898c8;
    font-size: 11px;
    font-family: 'Segoe UI';
    min-width: 30px;
}

/* Inputs */
#TextInput, #FontCombo, #ValueSpin {
    background: rgba(0,0,0,0.35);
    color: #dde0f5;
    border: 1px solid rgba(90,90,170,0.2);
    border-radius: 6px;
    padding: 5px 7px;
    font-size: 12px;
    font-family: 'Segoe UI';
    selection-background-color: #6c5ce7;
}
#TextInput:focus, #FontCombo:focus, #ValueSpin:focus {
    border-color: rgba(108,92,231,0.55);
}
#FontCombo { min-width: 120px; }
#FontCombo QAbstractItemView {
    background: #191935;
    color: #dde0f5;
    border: 1px solid rgba(90,90,170,0.25);
    selection-background-color: #6c5ce7;
}

/* Buttons */
#PrimaryButton {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
        stop:0 #6c5ce7, stop:1 #a855f7);
    color: white; border: none; border-radius: 8px;
    padding: 8px 14px; font-weight: 600; font-size: 12px;
}
#PrimaryButton:hover {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
        stop:0 #7c6cf7, stop:1 #b865ff);
}
#PrimaryButton:pressed {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
        stop:0 #5c4cd7, stop:1 #9845e7);
}

#SecondaryButton {
    background: rgba(255,255,255,0.06);
    color: #c0c0e0;
    border: 1px solid rgba(140,140,210,0.2);
    border-radius: 8px; padding: 8px 12px; font-size: 12px;
}
#SecondaryButton:hover {
    background: rgba(255,255,255,0.12);
    border-color: rgba(140,140,210,0.4); color: white;
}

#StyleToggleBtn {
    background: rgba(255,255,255,0.05);
    color: #8888bb;
    border: 1px solid rgba(100,100,180,0.2);
    border-radius: 6px;
}
#StyleToggleBtn:checked {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
        stop:0 #6c5ce7, stop:1 #a855f7);
    color: white; border-color: #6c5ce7;
}

#ToggleWordBtn {
    background: rgba(255,255,255,0.06);
    color: #9999bb;
    border: 1px solid rgba(100,100,180,0.2);
    border-radius: 5px; padding: 2px 12px;
    font-size: 11px;
}
#ToggleWordBtn:checked {
    background: rgba(108,92,231,0.5);
    color: white; border-color: #6c5ce7;
}

#RunBatchBtn {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
        stop:0 #00b894, stop:1 #00cec9);
    color: white; border: none; border-radius: 8px;
    padding: 8px 18px; font-weight: 700; font-size: 13px;
}
#RunBatchBtn:hover {
    background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
        stop:0 #00d9a3, stop:1 #00e0d9);
}
#RunBatchBtn:disabled {
    background: rgba(100,100,140,0.3);
    color: rgba(200,200,220,0.5);
}

/* Preview */
#PreviewWrapper { background: #090916; }
#PreviewTitle { color: #c0c0e0; padding: 2px 0; }
#InfoBar {
    background: rgba(18,18,38,0.9);
    border: 1px solid rgba(80,80,160,0.12);
    border-radius: 7px; margin-top: 6px;
}
#InfoLabel { color: #8888b5; font-size: 11px; }

/* Log Panel */
#LogPanel {
    background: rgba(8,8,20,0.95);
    color: #b0b0d0;
    border: 1px solid rgba(80,80,160,0.15);
    border-radius: 8px;
    padding: 8px;
    font-family: 'Consolas', 'Courier New', monospace;
    font-size: 10px;
}

/* Bottom */
#BottomBar {
    border-top: 1px solid rgba(90,90,170,0.12);
}

/* Scroll */
#ContentScroll { background: transparent; border: none; }
#ContentWidget { background: transparent; }

QScrollBar:vertical {
    background: rgba(0,0,0,0.2); width: 7px; border-radius: 3px;
}
QScrollBar::handle:vertical {
    background: rgba(108,92,231,0.4); border-radius: 3px; min-height: 28px;
}
QScrollBar::handle:vertical:hover { background: rgba(108,92,231,0.7); }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }

QSplitter::handle { background: rgba(90,90,170,0.15); }
QSplitter::handle:hover { background: rgba(108,92,231,0.5); }

/* SpinBox arrows */
QSpinBox::up-button, QDoubleSpinBox::up-button {
    background: rgba(108,92,231,0.2);
    border-left: 1px solid rgba(90,90,170,0.12);
    border-top-right-radius: 5px; width: 16px;
}
QSpinBox::up-button:hover, QDoubleSpinBox::up-button:hover {
    background: rgba(108,92,231,0.4);
}
QSpinBox::down-button, QDoubleSpinBox::down-button {
    background: rgba(108,92,231,0.2);
    border-left: 1px solid rgba(90,90,170,0.12);
    border-bottom-right-radius: 5px; width: 16px;
}
QSpinBox::down-button:hover, QDoubleSpinBox::down-button:hover {
    background: rgba(108,92,231,0.4);
}
QSpinBox::up-arrow, QDoubleSpinBox::up-arrow {
    image: none; width: 0; height: 0;
    border-left: 3px solid transparent;
    border-right: 3px solid transparent;
    border-bottom: 4px solid #9999cc;
}
QSpinBox::down-arrow, QDoubleSpinBox::down-arrow {
    image: none; width: 0; height: 0;
    border-left: 3px solid transparent;
    border-right: 3px solid transparent;
    border-top: 4px solid #9999cc;
}
"""


# ════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════
def main():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))

    if FONTS_DIR.exists():
        for f in FONTS_DIR.glob("*.ttf"):
            QFontDatabase.addApplicationFont(str(f))

    win = TextOverlayEditor()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
